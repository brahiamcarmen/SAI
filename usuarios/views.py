# -*- coding: utf-8 -*-
from django.shortcuts import render
from django.conf import settings
from django.utils import timezone
from django.core.exceptions import ObjectDoesNotExist
from django.views.generic.base import View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.shortcuts import render, get_object_or_404
from SAAL.models import Usuario, Tarifa, Credito, AsignacionBloque, Novedades
from SAAL.models import Poblacion, Facturas, Ciclo, EstadoCuenta, AcuerdosPago, FacturasConceptos
from SAAL.models import Vivienda, SolicitudGastos, Propietario, Medidores, Pqrs, RespuestasPqrs
from SAAL.models import CobroMatricula, Permisos, Pagos, Cierres, Acueducto, ValorMatricula, ConsumosMensual
from SAAL.models import Proveedor,Asignacion, Consumos, Conceptos, ConceptosFacturados, OrdenesTrabajo
from SAAL.forms import FormAgregarGasto, FormRegistroPqrs, RegistroVivienda
from SAAL.forms import AcueductoAForm, CobroMatriculaForm, CostoMForm, FormRespuestaPqrs, FormAsignarMedidor
from SAAL.forms import RegistroPropietario, TarifasForm, ModificaPropietario, FormRegistroCredito, FormRegistroProveedor
from SAAL.forms import CambioFormEstado, AcueductoForm, GastosForm, MedidoresForm, ModificaVivienda
from django.contrib import messages
from django.contrib.auth.models import User
from django.http import HttpResponseRedirect
from django.urls import reverse
from django.contrib import auth
from datetime import datetime
from datetime import timedelta
from django.http import HttpResponse
from openpyxl import Workbook
import openpyxl
import qrcode
from openpyxl.drawing.image import Image
from django.template.loader import get_template
from django.core.mail import EmailMultiAlternatives
from usuarios.ConectorPython import *
from django.db.models import Sum
from calendar import monthrange, month_name
import locale
from django.db.models.functions import ExtractYear

# Reemplaza estos valores con tus credenciales de Google Mail
username = 'sistemas.acueducto.caimalito@gmail.com'

#SECTORES
SECTOR1 = 'Pasonivel Viejo'
SECTOR2 = 'Pasonivel Destapada'
SECTOR3 = 'Caimalito Centro'
SECTOR4 = 'Barrio Nuevo'
SECTOR5 = '20 de julio'
SECTOR6 = 'Carbonera'
SECTOR7 = 'Hacienda'

# ESTADOS PRERIOS
ESTADOS1 = 'Operativo'
ESTADOS2 = 'Suspendido'
ESTADOS3 = 'Mantenimiento'
ESTADOS4 = 'Retirado'

#NOVEDADES
NOVEDAD1 = 'Cerrada'

# Tiempos de facturacion
DIASFACTURACION = 10
DIASPARASUSPENCION = 15
# permisos
CT = 'AC'
# TARIFA
TARIFA = 10000
# ------------
EF = 'Emitida'
EPC = 'Se registro propietario'
EPM = 'Se modifico propietario'
DESC = 'Null'
ECV = 'Se registro vivienda'
EMV = 'Se modifico vivienda'
# Tipo novedades
DES = 'Descuento'
ADI = 'Adicion'
# estados suspenciones
SA = 'Anulada'
SP = 'Pendiente'
SJ = 'Ejecutada'
TARIFASUSPENCION = 5000
# estados ciclos
EC = 'SIN PAGAR'
EC2 = 'PAGO'
EC3 = 'ANULADA'
# EstadosFacturas
FE = 'Emitida'
FV = 'Vencida'
FP = 'Paga'
FA = 'Anulada'
# Estadoscuentasdecobro
Estadocue = 'Mantenimiento'
# otros
REPORTESUSPEN = 'Suspedido'
# sectores
S1 = 'Pasonivel Viejo'
S2 = 'Pasonivel Destapada'
S3 = 'Caimalito Centro'
S4 = 'Barrio Nuevo'
S5 = '20 de julio'

# ESTADOS PRERIOS
E1 = 'Operativo'
E2 = 'Suspendido'
E3 = 'Retirado'
E4 = 'En construcion'

C1 = 'Ciclo 1'
C2 = 'Ciclo 2'
C3 = 'Ciclo 3'
C4 = 'Ciclo 4'

DESCOBRO = 'Concepto matricula'
COBROCONSUMO = 'Cobro por consumo'
ESTCOBRO = 'Pendiente'
ESTCOBRO2 = 'Pago'
ESTADO1 = "Pendiente"
ESTADO2 = "Aprobada"
ESTADO3 = "Anulada"

class Inicio(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/inicio.html'
    form_class = AcueductoForm

    def get(self, request):
        try:
            version = open('static/serial/Version.txt', 'r')
            versionp = version.read()
            version2 = open('static/serial/NombreProyecto.txt', 'r')
            nombreproyecto = version2.read()
            version3 = open('static/serial/NombreProyectoL.txt', 'r')
            nombreproyectol = version3.read()
            hoy = datetime.now()
            usuario = Usuario.objects.get(usuid=request.user.pk)
            idacueducto = usuario.IdAcueducto
            tipousuario = usuario.TipoUsuario
            acueducto = Acueducto.objects.get(IdAcueducto=idacueducto)
            nombreacueducto = acueducto.Sigla
            logo = acueducto.logo
            usuarios = Usuario.objects.filter(IdAcueducto=idacueducto)
            novedades = Novedades.objects.filter(IdAcueducto=idacueducto).order_by("-IdNovedad")[:6]
            # mensualidades:
            fechaexp = (datetime.today())
            ciclo = fechaexp.month
            ano1 = fechaexp.year
            # Los argumentos serán: Año, Mes, Día, Hora, Minutos, Segundos, Milisegundos.
            new_date = datetime(ano1, ciclo, 1, 1, 00, 00, 00000)
            new_date2 = datetime(ano1, ciclo, 28, 23, 59, 59, 00000)
            pagos2 = Pagos.objects.filter(FechaPago__gte=new_date, FechaPago__lte=new_date2, IdAcueducto=idacueducto).all()
            pago0 = 0
            for i in pagos2:
                valor = i.ValorPago
                pago0 += int(valor)

            predios = Vivienda.objects.filter(EstadoServicio='Operativo', IdAcueducto=idacueducto).count()
            recaudot = int(predios) * 10000
            if pago0 ==0:
                porcentaje =0
            else:
                porcentaje = pago0 / recaudot * 100
            # mensualidades:
            ciclo2 = fechaexp.month - 1
            ano2 = fechaexp.year
            # Los argumentos serán: Año, Mes, Día, Hora, Minutos, Segundos, Milisegundos.
            new_date3 = datetime(ano2, ciclo2, 1, 1, 00, 00, 00000)
            new_date4 = datetime(ano2, ciclo2, 28, 23, 59, 59, 00000)
            factuasemi = Facturas.objects.filter(FechaExpe__gte=new_date3, FechaExpe__lte=new_date4, IdAcueducto=idacueducto).count()
            pagos3 = Pagos.objects.filter(FechaPago__gte=new_date, FechaPago__lte=new_date2, IdAcueducto=idacueducto).count()

            if pagos3 ==0:
                promedio = 0
                promtarifa = 0
                contador = 0
            else:
                promedio = pago0 / pagos3
                promtarifa = promedio / 10000 * 100
                contador = pagos3 / factuasemi * 100

            viviendas = Vivienda.objects.filter(EstadoServicio='Operativo',IdAcueducto=idacueducto) | Vivienda.objects.filter(EstadoServicio='Mantenimiento',IdAcueducto=idacueducto)| Vivienda.objects.filter(EstadoServicio='Suspendido',IdAcueducto=idacueducto)
            personas = 0
            for i in viviendas:
                valor = int(i.CantHabitantes)
                personas += valor

            suscriptoresactivos = Vivienda.objects.filter(EstadoServicio='Operativo',IdAcueducto=idacueducto) | Vivienda.objects.filter(EstadoServicio='Suspendido',IdAcueducto=idacueducto)
            suscriptores = 0
            for i in suscriptoresactivos:
                valor = 1
                suscriptores += valor

            contarasig = Asignacion.objects.filter(Estado='Operativo',IdAcueducto=idacueducto).count()
            vapo = Consumos.objects.filter(IdAcueducto=idacueducto).aggregate(Consumo=Sum('Consumo'))
            suma8 = vapo['Consumo']

            operativos = Vivienda.objects.filter(EstadoServicio='Operativo', IdAcueducto=idacueducto).count()
            suspendidos = Vivienda.objects.filter(EstadoServicio='Suspendido', IdAcueducto=idacueducto).count()
            retirados = Vivienda.objects.filter(EstadoServicio='Retirado', IdAcueducto=idacueducto).count()

            #estratos
            estr1 = Vivienda.objects.filter(Estrato='1', IdAcueducto=idacueducto).count()
            estr2 = Vivienda.objects.filter(Estrato='2', IdAcueducto=idacueducto).count()
            estr3 = Vivienda.objects.filter(Estrato='3', IdAcueducto=idacueducto).count()
            estr4 = Vivienda.objects.filter(Estrato='4', IdAcueducto=idacueducto).count()
            estr5 = Vivienda.objects.filter(Estrato='5', IdAcueducto=idacueducto).count()
            comercial = Vivienda.objects.filter(Estrato='Comercial', IdAcueducto=idacueducto).count()
            industrial = Vivienda.objects.filter(Estrato='Industrial', IdAcueducto=idacueducto).count()
            especial = Vivienda.objects.filter(Estrato='Especial', IdAcueducto=idacueducto).count()
            oficial = Vivienda.objects.filter(Estrato='Oficial', IdAcueducto=idacueducto).count()

            asignado = Vivienda.objects.filter(EstadoServicio='Operativo', IdAcueducto=idacueducto, TipoRecaudo='Medicion').count()
            cant = operativos + suspendidos
            prediossinmedidor = cant - asignado

            vapo = Consumos.objects.filter(IdAcueducto=idacueducto).aggregate(Consumo=Sum('Consumo'))
            sumaconsumo = vapo['Consumo']


            ventas = ConsumosMensual.objects.all().values('mes', 'Consumo','CantMedidores')[:12]

            fechas = [e['mes'] for e in ventas]
            totales = [e['Consumo'] for e in ventas]
            cantmedi = [e['CantMedidores'] for e in ventas]

            messages.success(request, "¡Bienvenido/a al sistema!")
            return render(request,
                          self.template_name, {'tipousuario': tipousuario, 'nombreproyecto': nombreproyecto,'sinmedidor':prediossinmedidor, 'asignado':asignado,
                                               'nombreproyectol': nombreproyectol, 'acueducto': nombreacueducto,'consumos':sumaconsumo,
                                               'versionp': versionp,'personas':personas,'nit':idacueducto,
                                               'novedades': novedades, 'pagos': pago0,'suscriptores':suscriptores,
                                               'porcentaje': int(porcentaje), 'contador': int(contador),
                                               'facturaspagas': pagos3, 'promedio': int(promedio),'logo':logo,
                                               'promtarifa': int(promtarifa),'contarasig': contarasig, 'suma8': suma8,
                                               'foto': usuario.fotoUsuario,'usuarios': usuarios,'celular': usuario.celular,
                                               'departamento': usuario.Departamento,'cargo': usuario.Cargo, 'fechac': usuario.FechaCreacion,
                                               'ultimo': hoy, 'retirados': retirados,'suspendidos': suspendidos,'operativos':operativos,
                                               'uno': estr1,'dos':estr2,'tres':estr3,'cuatro':estr4,'cinco':estr5,'especial':especial,'industrial':industrial,
                                               'oficial':oficial,'comercial':comercial, 'labels': fechas,'data': totales,'cantmedidores':cantmedi})
        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class ListaViviendas(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/listaviviendas.html'

    def get(self, request):
        # Obtener el usuario actual
        usuario = get_object_or_404(Usuario, usuid=request.user.pk)

        # Obtener el acueducto asociado
        acueducto = get_object_or_404(Acueducto, IdAcueducto=usuario.IdAcueducto.pk)

        # Filtrar las viviendas por el acueducto
        listaviviendas = Vivienda.objects.filter(IdAcueducto=acueducto.pk)

        # Renderizar la plantilla con las viviendas
        return render(request, self.template_name, {'viviendas': listaviviendas})

class ListaPropietarios(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/listapropietarios.html'

    def get(self, request):
        try:

            datos = Usuario.objects.get(usuid=request.user.pk)
            dr = datos.IdAcueducto
            usuarios = Propietario.objects.filter(IdAcueducto=dr)
            tipousuario = True
            if tipousuario is True:
                return render(request, self.template_name, {'usuarios': usuarios})
            else:
                messages.add_message(request, messages.ERROR, 'Su usuario no tiene los permisos de acceso a esta seccion')
                return HttpResponseRedirect(reverse('usuarios:inicio'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class AgregarPropietario(LoginRequiredMixin, View):
    login_url = '/'
    form_class = RegistroPropietario
    template_name = 'usuarios/registropropietario.html'

    def get(self, request):
        try:
            form = self.form_class()
            usuario = Usuario.objects.get(usuid=request.user.pk)
            tipousuario = True
            if tipousuario is True:
                return render(request, self.template_name,
                              {
                                  'form': form,
                              })
            else:
                messages.add_message(request, messages.ERROR, 'Su usuario no tiene los permisos de acceso a esta seccion')
                return HttpResponseRedirect(reverse('usuarios:inicio'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request):
        try:
            usuario = Usuario.objects.get(usuid=request.user.pk)
            idacueducto = usuario.IdAcueducto
            acueducto = Acueducto.objects.get(IdAcueducto=idacueducto)
            idpropietario = request.POST.get("IdPropietario")
            nombres = request.POST.get("Nombres")
            apellidos = request.POST.get("Apellidos")
            notelefono = request.POST.get("NoTelefono")
            email = request.POST.get("Email")
            idpoblacion = request.POST.get("IdPoblacion")
            poblacion = Poblacion.objects.get(IdPoblacion=idpoblacion)

            validarpro = Propietario.objects.filter(IdPropietario=idpropietario).exists()
            if validarpro is True:
                messages.add_message(request, messages.ERROR, 'El usuario ya existe')
                return HttpResponseRedirect(reverse('usuarios:agregarpropietario'))

            else:
                if int(idpropietario) <= 400000:
                    messages.add_message(request, messages.ERROR, 'El numero de identificacion del propietario no es valido')
                    return HttpResponseRedirect(reverse('usuarios:agregarpropietario'))

                else:
                    propietario = Propietario(IdPropietario=idpropietario, Nombres=nombres, Apellidos=apellidos,
                                              NoTelefono=notelefono, Email=email, IdPoblacion=poblacion, IdAcueducto=acueducto)

                    if propietario is not None:
                        propietario.save()
                        messages.add_message(request, messages.INFO, 'el propietario se agrego correctamente')
                        return HttpResponseRedirect(reverse('usuarios:listapropietarios'))

                    else:
                        messages.add_message(request, messages.ERROR, 'El propietario no se pudo agregar')
                        return HttpResponseRedirect(reverse('usuarios:agregarpropietario'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class ModificarPropietario(LoginRequiredMixin, View):
    login_url = '/'
    form_class = ModificaPropietario
    template_name = 'usuarios/modificarpropietario.html'

    def get(self, request, IdPropietario):
        try:
            datospropietario = Propietario.objects.get(IdPropietario=IdPropietario)
            form = self.form_class(instance=datospropietario)
            usuario = Usuario.objects.get(usuid=request.user.pk)
            tipousuario = Permisos.objects.filter(usuid=usuario, TipoPermiso='NOACMT').exists()
            if tipousuario is False:
                return render(request, self.template_name, {'form': form})
            else:
                messages.add_message(request, messages.ERROR, 'Su usuario no tiene los permiso de acceso a '
                                                              'esta seccion')
                return HttpResponseRedirect(reverse('usuarios:inicio'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request, IdPropietario):
        try:
            datospropietario = Propietario.objects.get(IdPropietario=IdPropietario)
            form = self.form_class(request.user, request.POST, instance=datospropietario)

            if form.is_valid():
                form.save()
                messages.add_message(request, messages.INFO, 'la informacion del propietario se modifico correctamente')
                return HttpResponseRedirect(reverse('usuarios:listapropietarios'))

            else:
                messages.add_message(request, messages.ERROR, 'No se puedo modificar la informacion')
                return HttpResponseRedirect(reverse('usuarios:listapropietarios'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class ModificarVivienda(LoginRequiredMixin, View):
    login_url = '/'
    form_class = ModificaVivienda
    template_name = 'usuarios/modificarvivienda.html'

    def get(self, request, idvivienda):
        try:
            idvivienda = str(idvivienda)
            datosvivienda = Vivienda.objects.get(IdVivienda=idvivienda)
            form = self.form_class(instance=datosvivienda)
            usuario = Usuario.objects.get(usuid=request.user.pk)
            tipousuario = False
            if tipousuario is False:
                return render(request, self.template_name, {'form': form, 'matricula': idvivienda})

            else:
                messages.add_message(request, messages.ERROR, 'Su usuario no tiene los permiso de '
                                                              'acceso a esta seccion')
                return HttpResponseRedirect(reverse('usuarios:inicio'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request, idvivienda):
        try:
            datosvivienda = Vivienda.objects.get(IdVivienda=idvivienda)
            form = self.form_class(request.user, request.POST, instance=datosvivienda)

            if form.is_valid():
                form.save()
                messages.add_message(request, messages.INFO, 'la informacion de la vivienda se modifico correctamente')
                return HttpResponseRedirect(reverse('usuarios:listaviviendas'))

            else:
                messages.add_message(request, messages.ERROR, 'No se puedo modificar la informacion')
                return HttpResponseRedirect(reverse('usuarios:listaviviendas'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class VisualizarPropietario(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/verpropietario.html'

    def get(self, request, idpropietario):
        try:
            datospropietario = Propietario.objects.get(IdPropietario=idpropietario)
            viviendas = Vivienda.objects.filter(IdPropietario=idpropietario)
            return render(request, self.template_name,
                          {
                              'viviendas': viviendas,
                              'IdPropietario': datospropietario.IdPropietario,
                              'Nombres': datospropietario.Nombres,
                              'Apellidos': datospropietario.Apellidos,
                              'NoTelefono': datospropietario.NoTelefono,
                              'Email': datospropietario.Email,
                              'IdPoblacion': datospropietario.IdPoblacion
                          })

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class VisualizarVivienda(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/vervivienda.html'

    def get(self, request, idvivienda):
        try:
            vivienda = Vivienda.objects.get(IdVivienda=idvivienda)

            # Consultas de uso frecuente y optimización de conteo
            viviendainfo = Vivienda.objects.filter(IdVivienda=idvivienda)
            estados = EstadoCuenta.objects.filter(IdVivienda=idvivienda)
            cobromatricula = CobroMatricula.objects.filter(IdVivienda=idvivienda, Estado=ESTCOBRO)
            pagos = Pagos.objects.filter(IdVivienda=idvivienda)
            facturas = Facturas.objects.filter(IdVivienda=idvivienda).order_by("-IdFactura")
            nofacturas = facturas.count()
            facturasemi = facturas[:1]
            vafacemi = Facturas.objects.filter(IdVivienda=idvivienda, Estado=FE).exists()
            matriculas = CobroMatricula.objects.filter(IdVivienda=idvivienda)
            matriculas2 = matriculas.exists()
            suspenciones = OrdenesTrabajo.objects.filter(IdVivienda=idvivienda).order_by("-IdOrden")
            filtrosuspenciones = suspenciones.filter(TipoNovedad='Suspension', Estado='Pendiente').exists()
            validarretiro = Novedades.objects.filter(matricula=idvivienda, TipoNovedad='Retiro').exists()
            novretiro = Novedades.objects.filter(matricula=idvivienda, TipoNovedad='Retiro')
            conceptos = Conceptos.objects.filter(IdVivienda=idvivienda).order_by("-IdRegistro")
            conceptosfacturados = ConceptosFacturados.objects.filter(IdVivienda=idvivienda).order_by("-IdRegistro")
            consumos = Consumos.objects.filter(IdVivienda=idvivienda).order_by("-IdRegistro")
            reparaciones = 0

            # Calcular suma de conceptos sin facturar y facturados pendientes
            suma2 = Conceptos.objects.filter(Estado='Sin facturar', IdVivienda=idvivienda).aggregate(Sum('Valor'))['Valor__sum'] or 0
            suma = ConceptosFacturados.objects.filter(Estado='Pendiente', IdVivienda=idvivienda).aggregate(Sum('Total'))['Total__sum'] or 0
            total = suma + suma2

            # Cálculo de cuotas de matrícula
            matri = cobromatricula.aggregate(Sum('Cuota'))['Cuota__sum'] or 0

            # Facturas pagadas
            lista = [k.IdFactura for k in facturas if Pagos.objects.filter(IdFactura=k.IdFactura).exists()]

            # Total pagado
            pagado = pagos.aggregate(Sum('ValorPago'))['ValorPago__sum'] or 0

            # Consumos de los últimos 6 meses
            ventas = Consumos.objects.filter(IdVivienda=idvivienda).values('mes', 'Consumo').order_by("-IdRegistro")[:6]
            meses = [e['mes'] for e in ventas]
            consumosm = [e['Consumo'] for e in ventas]

            # Consultar si hay asignaciones y acuerdos de pago
            asignado = Asignacion.objects.filter(IdVivienda=idvivienda, Estado='Operativo').exists()
            novedades = Novedades.objects.filter(matricula=idvivienda)
            acuerdospago = AcuerdosPago.objects.filter(IdVivienda=idvivienda)

            # Renderizar la plantilla
            return render(request, self.template_name, {
                'pagado': pagado, 'nofac': nofacturas, 'asignado': asignado, 'labels': meses, 'data': consumosm,
                'lista': lista, 'facturas': facturas, 'cobromatricula': cobromatricula, 'suspenciones': suspenciones,
                'facturasemi': facturasemi, 'matriculas': matriculas, 'total': total, 'aportes': suma,
                'aportes2': suma2, 'direccion': vivienda.Direccion, 'casa': vivienda.NumeroCasa, 'piso': vivienda.Piso,
                'matricula': vivienda.IdVivienda, 'tipo': vivienda.TipoInstalacion, 'acuerdos': acuerdospago,
                'estrato': vivienda.Estrato, 'tipop': vivienda.InfoInstalacion, 'estado': vivienda.EstadoServicio,
                'propietario': vivienda.IdPropietario, 'fichacatastral': vivienda.FichaCastral,
                'estados': estados, 'pagos': pagos, 'fecha': timezone.now(),
                'ultimopago': pagos.order_by("-IdPago")[:1], 'vafacemi': vafacemi, 'viviendainfo': viviendainfo,
                'cobromatricula1': matri, 'novedades': novedades, 'repaciones': reparaciones, 'conceptos': conceptos, 'conceptosfacturados': conceptosfacturados,
                'consumos': consumos, 'filtro': filtrosuspenciones, 'contpagos': pagos.count(), 'vmatri': matriculas2,
                'novedadr': validarretiro,'novretiro': novretiro,
            })

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class AgregarVivienda(LoginRequiredMixin, View):
    login_url = '/'
    form_class = RegistroVivienda
    vizualizarv = VisualizarVivienda
    template_name = 'usuarios/registrovivienda.html'

    def get(self, request, idbloque):
        try:
            form = self.form_class()
            matricula = AsignacionBloque.objects.get(IdBloque=idbloque)
            asignada = matricula.Matricula
            usuario = Usuario.objects.get(usuid=request.user.pk)
            tipousuario = Permisos.objects.filter(usuid=usuario, TipoPermiso='AIP').exists()
            if tipousuario is True:
                return render(request, self.template_name,
                              {
                                  'form': form,
                                  'asignada': asignada
                              })
            else:
                messages.add_message(request, messages.ERROR, 'Su usuario no tiene los permisos de acceso a esta '
                                                              'seccion')
                return HttpResponseRedirect(reverse('usuarios:inicio'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request, idbloque):
        try:
            matricula = AsignacionBloque.objects.get(IdBloque=idbloque)
            asignada = matricula.Matricula
            idvivienda = asignada
            direccion = request.POST.get("Direccion")
            numerocasa = request.POST.get("NumeroCasa")
            piso = request.POST.get("Piso")
            ciclo = request.POST.get("Ciclo")
            tipoinstalacion = request.POST.get("TipoInstalacion")
            estrato = request.POST.get("Estrato")
            estadoservicio = request.POST.get("EstadoServicio")
            idpropietario = request.POST.get("IdPropietario")
            matricula = request.POST.get("MatriculaAnt")
            infoinstalacion = request.POST.get("InfoInstalacion")
            profacometida = request.POST.get("ProfAcometida")
            canthabitantes = request.POST.get("CantHabitantes")
            fichacatastral = request.POST.get("FichaCastral")
            diametro = request.POST.get("Diametro")
            datos = Usuario.objects.get(usuid=request.user.pk)
            dr = datos.IdAcueducto
            acueducto = Acueducto.objects.get(IdAcueducto=dr.pk)
            propietario = Propietario.objects.get(IdPropietario=idpropietario)
            validarvi = Vivienda.objects.filter(IdVivienda=idvivienda).exists()
            if validarvi is True:
                messages.add_message(request, messages.ERROR, 'la Vivienda ya existe')
                return HttpResponseRedirect(reverse('usuarios:matriculas'))

            else:
                vivienda = Vivienda(IdVivienda=idvivienda, Direccion=direccion, NumeroCasa=numerocasa, Piso=piso,
                                    Ciclo=ciclo, TipoInstalacion=tipoinstalacion, Estrato=estrato,
                                    EstadoServicio=estadoservicio, IdPropietario=propietario, MatriculaAnt=matricula,
                                    InfoInstalacion=infoinstalacion, ProfAcometida=profacometida,
                                    CantHabitantes=canthabitantes, IdAcueducto=acueducto,
                                    FichaCastral=fichacatastral, Diametro=diametro, usuid=datos.usuid)
                vivienda.save()
                estadocuenta = EstadoCuenta(Valor=0, IdVivienda=vivienda, Estado='Operativo', Descripcion=COBROCONSUMO)
                estadocuenta.save()
                matriculas = AsignacionBloque.objects.get(IdBloque=idbloque)
                matriculas.Estado = 'Asignada'
                matriculas.Estadocuenta = estadocuenta.IdEstadoCuenta
                matriculas.save()
                ver = self.vizualizarv()
                messages.add_message(request, messages.INFO, 'La informacion del predio se agrego correctamente')
                ejercutar = ver.get(request, idvivienda)
                return ejercutar

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class Estadoscuenta(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/estadosdecuenta.html'

    def get(self, request):
        try:
            usuario = Usuario.objects.get(usuid=request.user.pk)
            idacueducto = usuario.IdAcueducto
            totalcuentas = EstadoCuenta.objects.all().count()

            totalfac = Facturas.objects.filter(IdAcueducto=idacueducto).count()
            facemi2 = Facturas.objects.filter(Estado=FE,IdAcueducto=idacueducto).count()
            facven = Facturas.objects.filter(Estado=FV,IdAcueducto=idacueducto).count()
            facpg = Facturas.objects.filter(Estado=FP,IdAcueducto=idacueducto).count()
            facanu = Facturas.objects.filter(Estado=FA,IdAcueducto=idacueducto).count()
            acuerdospagos = AcuerdosPago.objects.filter(IdAcueducto=idacueducto).order_by("-IdAcuerdo")
            acuerdospago = AcuerdosPago.objects.filter(IdAcueducto=idacueducto).count()
            acuerpen = AcuerdosPago.objects.filter(IdAcueducto=idacueducto, Estado='Pendiente').count()
            acuerpag = AcuerdosPago.objects.filter(IdAcueducto=idacueducto, Estado='Pago').count()

            facturasvalor = Facturas.objects.filter(Estado='Emitida',IdAcueducto=idacueducto).aggregate(Total=Sum('Total'))
            total = facturasvalor['Total']

            vapo = ConceptosFacturados.objects.filter(Estado='Pendiente',IdAcueducto=idacueducto).aggregate(Total=Sum('Total'))
            sumatotal = vapo['Total']

            resultados = ConceptosFacturados.objects.filter(Estado='Pendiente', IdAcueducto=idacueducto).aggregate(
                AporteFijo=Sum('AporteFijo'),
                Basico=Sum('Basico'),
                Complementario=Sum('Complementario'),
                Suspencion=Sum('Suspencion'),
                Reconexion=Sum('Reconexion'),
                Recargo=Sum('Recargo'),
                Subsidio=Sum('Subsidio'),
                AcuerdoPago=Sum('AcuerdoPago'),
                SaldoAnterior=Sum('SaldoAnterior')
            )

            # Extraer los resultados de cada sumatoria
            sum1 = resultados['AporteFijo'] or 0
            sum2 = resultados['Basico'] or 0
            sum3 = resultados['Complementario'] or 0
            sum4 = resultados['Suspencion'] or 0
            sum5 = resultados['Reconexion'] or 0
            sum6 = resultados['Recargo'] or 0
            sum7 = resultados['Subsidio'] or 0
            sum8 = resultados['AcuerdoPago'] or 0
            sum9 = resultados['SaldoAnterior'] or 0

            sinfac = Conceptos.objects.filter(Estado='Sin facturar').aggregate(Valor=Sum('Valor'))
            sum10 = sinfac['Valor']

            totalconceptos = sum1 + sum2 + sum3 + sum4 + sum5 + sum6 + sum8 + sum9 + sum10
            tipousuario = True
            if tipousuario is True:
                return render(request, self.template_name,
                              {
                                  'totalcuentascobro': totalcuentas, 'total': total, 'vapo': sumatotal,'acuerdos':acuerdospago,
                                  'totalfac':totalfac,  'facven':facven, 'facanu': facanu, 'facpg':facpg, 'facemi': facemi2,'acuerdospagos': acuerdospagos, 'totales': totalconceptos,
                                  'apf': sum1,'bss': sum2, 'comp': sum3,'sus': sum4,'re': sum5,'recar': sum6, 'sub':sum7, 'acuer': sum8,'anter':sum9,'sinfac':sum10,
                                  'acuerpa': acuerpag, 'acuerpen': acuerpen
                              })
            else:
                messages.add_message(request, messages.ERROR, 'Su usuario no tiene los permiso de acceso a esta seccion')
                return HttpResponseRedirect(reverse('usuarios:inicio'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request):
        try:
            matricula = request.POST.get("Matricula")
            valor = request.POST.get("Valor")
            cuotas = request.POST.get("Cuotas")
            descripcion = request.POST.get("Descripcion")

            usuario = Usuario.objects.get(usuid=request.user.pk)
            idacueducto = usuario.IdAcueducto
            acueducto = Acueducto.objects.get(IdAcueducto=idacueducto)

            cuota = int(valor) / int(cuotas)
            viviendaw = Vivienda.objects.filter(IdVivienda=matricula).exists()
            if viviendaw is True:
                vivienda =Vivienda.objects.get(IdVivienda=matricula)
                acuerdo = AcuerdosPago(Tipo='Financiacion',Descripcion=descripcion, IdVivienda=vivienda, Estado='Pendiente',
                                       Valor=valor,CantCuotas=cuotas,CuotasPendientes=cuotas,ValorPendiente=valor,Cuota=cuota,
                                       IdAcueducto=acueducto)
                acuerdo.save()
                messages.add_message(request, messages.INFO,'Se registro la financiacion correctamente')
                return HttpResponseRedirect(reverse('usuarios:estadoscuenta'))

            else:
                messages.add_message(request, messages.ERROR, 'La matricula ingresada no esta registrada en el sistema o ya tiene un acuerdo de pago')
                return HttpResponseRedirect(reverse('usuarios:estadoscuenta'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

class Busquedas(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/busquedas.html'
    propietario = VisualizarPropietario
    predio = VisualizarVivienda

    def get(self, request):
        try:
            usuario = Usuario.objects.get(usuid=request.user.pk)
            idacueducto = usuario.IdAcueducto
            tipo = request.GET.get("tipo")
            identificacion = request.GET.get("identificacion")

            if tipo == "Cedula de ciudadania":
                titular = Propietario.objects.filter(IdPropietario=identificacion, IdAcueducto=idacueducto).exists()
                if titular is True:
                    idpropietario = identificacion
                    ver = self.propietario()
                    ejercutar = ver.get(request, idpropietario)
                    return ejercutar

                else:
                    messages.add_message(request, messages.ERROR, 'Informacion del titular no encontrada')
                    return HttpResponseRedirect(reverse('usuarios:busquedas'))

            elif tipo == "Numero de matricula" and identificacion is not None:
                predio = Vivienda.objects.filter(IdVivienda=identificacion, IdAcueducto=idacueducto).exists()
                if predio is True:
                    idvivienda = identificacion
                    ver = self.predio()
                    ejecutar = ver.get(request, idvivienda)
                    return ejecutar

                else:
                    messages.add_message(request, messages.ERROR, 'Informacion del predio no encontrada')
                    return HttpResponseRedirect(reverse('usuarios:inicio'))

            else:
                messages.add_message(request, messages.WARNING, 'Informacion incompleta')
                return HttpResponseRedirect(reverse('usuarios:inicio'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class ControlPresupuestal(LoginRequiredMixin, View):
    login_url = '/'
    form_class = GastosForm
    template_name = 'usuarios/gastos.html'

    def get(self, request):
        try:
            locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
            usuario = Usuario.objects.get(usuid=request.user.pk)
            idacueducto = usuario.IdAcueducto
            acueducto = Acueducto.objects.get(IdAcueducto=idacueducto)
            solicitudesgastos = SolicitudGastos.objects.filter(Estado=ESTADO1,IdAcueducto=idacueducto)
            form = self.form_class()
            contador = SolicitudGastos.objects.all().count()
            contadorpen = SolicitudGastos.objects.filter(Estado=ESTADO1,IdAcueducto=idacueducto).count()
            contadorapro = SolicitudGastos.objects.filter(Estado=ESTADO2,IdAcueducto=idacueducto).count()
            contadoranu = SolicitudGastos.objects.filter(Estado=ESTADO3,IdAcueducto=idacueducto).count()
            aprobado = SolicitudGastos.objects.filter(Estado=ESTADO2,IdAcueducto=idacueducto)

            credito = Credito.objects.filter(Estado='Vigente',IdAcueducto=idacueducto)
            pagos = Pagos.objects.filter(IdAcueducto=idacueducto)
            viviendasope = Vivienda.objects.filter(EstadoServicio=E1,IdAcueducto=idacueducto).count()
            tarifa = acueducto.IdTarifa.Valor

            totalpormes = int(int(viviendasope) * int(tarifa))
            pago = 0
            for i in pagos:
                valor = i.ValorPago
                pago += int(valor)

            suma2 = 0
            for i in aprobado:
                valor = int(i.Valor)
                suma2 += valor

            suma8 = 0
            for i in credito:
                valor = int(i.ValorPendiente)
                suma8 += valor

            totalingresos = pago
            gastos = int(suma2)
            presupuesto = totalingresos - gastos
            # mensualidades:
            fechaexp = (datetime.today())
            ciclo = fechaexp.month
            ano1 = fechaexp.year
            # Los argumentos serán: Año, Mes, Día, Hora, Minutos, Segundos, Milisegundos.
            new_date = datetime(ano1, ciclo, 1, 1, 00, 00, 00000)
            new_date2 = datetime(ano1, ciclo, 30, 23, 59, 59, 00000)
            pagos2 = Pagos.objects.filter(FechaPago__gte=new_date, FechaPago__lte=new_date2,IdAcueducto=idacueducto)
            gastosaprobados = SolicitudGastos.objects.filter(Fecha__gte=new_date, Fecha__lte=new_date2,
                                                             Estado=ESTADO2,IdAcueducto=idacueducto)

            gasto4 = 0
            for i in gastosaprobados:
                valor = int(i.Valor)
                gasto4 += valor

            pago0 = 0
            for i in pagos2:
                valor = i.ValorPago
                pago0 += int(valor)

            if totalpormes == 0:
                porcentaje = 0
            else:
                porcentaje = int(pago0 / totalpormes * 100)

            # Diccionario para almacenar los totales por mes
            datos_por_mes = {}
            # Totales acumulados para todo el año
            total_ingresos_anual = 0
            total_egresos_anual = 0

            # Iterar por cada mes del año
            for ciclo in range(1, 13):  # De enero (1) a diciembre (12)
                # Determinar el último día del mes dinámicamente
                ultimo_dia = monthrange(ano1, ciclo)[1]

                # Fechas de inicio y fin del mes
                fecha_inicio = datetime(ano1, ciclo, 1, 1, 00, 00, 00000)
                fecha_fin = datetime(ano1, ciclo, ultimo_dia, 23, 59, 59)

                # Consultar los pagos del mes actual
                pagos_mes = Pagos.objects.filter(
                    FechaPago__gte=fecha_inicio,
                    FechaPago__lte=fecha_fin,
                    IdAcueducto=idacueducto
                )
                gastos_mes = SolicitudGastos.objects.filter(
                    Fecha__gte=fecha_inicio,
                    Fecha__lte=fecha_fin,
                    Estado='Aprobada',
                    IdAcueducto=idacueducto
                )

                # Sumar el total de los pagos y gastos
                total_pagado_mes = sum(int(pago.ValorPago) for pago in pagos_mes)
                total_gastos_mes = sum(int(gasto.Valor) for gasto in gastos_mes)
                diferencia_mes = total_pagado_mes - total_gastos_mes
                total_ingresos_anual += total_pagado_mes
                total_egresos_anual += total_gastos_mes

            # Almacenar el total en el diccionario
                datos_por_mes[month_name[ciclo].capitalize()] = {
                    'total_ingresos': total_pagado_mes,
                    'total_egresos': total_gastos_mes,
                    'diferencia': diferencia_mes,
                }

            anos = [2021, 2022, 2023, 2024, 2025]
            datos_por_ano = {}

            for ano in anos:
                # Obtener pagos y gastos del año completo
                fecha_inicio = datetime(ano, 1, 1, 0, 0, 0)
                fecha_fin = datetime(ano, 12, 31, 23, 59, 59)

                # Consultar pagos y gastos del año
                pagos_ano = Pagos.objects.filter(
                    FechaPago__gte=fecha_inicio,
                    FechaPago__lte=fecha_fin,
                    IdAcueducto=idacueducto
                )
                gastos_ano = SolicitudGastos.objects.filter(
                    Fecha__gte=fecha_inicio,
                    Fecha__lte=fecha_fin,
                    Estado='Aprobada',
                    IdAcueducto=idacueducto
                )

                # Calcular totales
                total_ingresos_anio = sum(int(pago.ValorPago) for pago in pagos_ano)
                total_egresos_anio = sum(int(gasto.Valor) for gasto in gastos_ano)

                # Guardar resultados del año
                datos_por_ano[ano] = {
                    'total_ingresos_anio': total_ingresos_anio,
                    'total_egresos_anio': total_egresos_anio,
                    'diferencia_anio': total_ingresos_anio - total_egresos_anio,
                }

            tipousuario = True
            if tipousuario is True:
                return render(request, self.template_name, {
                    'porcentaje': porcentaje,
                    'form': form,
                    'solicitudesgastos': solicitudesgastos,
                    'contador': contador,
                    'contadorp': contadorpen,
                    'contadora': contadorapro,
                    'contadoranu': contadoranu,
                    'gastos': gastos,
                    'pago': int(totalingresos),
                    'presupuesto': presupuesto,
                    'ingresomensual': pago0,
                    'gastosmensuales': gasto4,
                    'credito': suma8,
                    'datos_por_mes': datos_por_mes,
                    'anio':ano1,
                    'total_ingresos_anual': total_ingresos_anual,
                    'total_egresos_anual': total_egresos_anual,
                    'datos_por_ano': datos_por_ano,
                })
            else:
                messages.add_message(request, messages.ERROR,
                                     'Su usuario no tiene los permisos de acceso a esta seccion')
                return HttpResponseRedirect(reverse('usuarios:inicio'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class GenerarGasto(LoginRequiredMixin, View):
    login_url = '/'
    form_class = FormAgregarGasto
    template_name = 'usuarios/generargasto.html'

    def get(self, request):
        try:
            form = self.form_class()
            return render(request, self.template_name, {
                'form': form,
            })

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request):
        try:
            usuario = Usuario.objects.get(usuid=request.user.pk)
            idacueducto = usuario.IdAcueducto
            acueducto = Acueducto.objects.get(IdAcueducto=idacueducto)
            area = request.POST.get("AreaResponsable")
            tiposolicitud = request.POST.get("TipoSolicitud")
            valor = request.POST.get("Valor")
            numerofactura = request.POST.get("NumeroFactura")
            descripcion = request.POST.get("Descripcion")
            proveedor = request.POST.get("IdProveedor")
            consulp = Proveedor.objects.get(IdProvedor=proveedor)
            usuario = Usuario.objects.get(usuid=request.user.pk)

            if area and numerofactura and tiposolicitud and valor and descripcion is not None:
                solicitud = SolicitudGastos(IdUsuario=usuario, Descripcion=descripcion,
                                            TipoSolicitud=tiposolicitud, Valor=valor,
                                            Estado=ESTADO1, AreaResponsable=area, NumeroFactura=numerofactura,
                                            IdProveedor=consulp, IdAcueducto=acueducto)
                solicitud.save()
                messages.add_message(request, messages.INFO, 'la solicitud se registro correctamente')
                return HttpResponseRedirect(reverse('usuarios:controlpresupuestal'))

            else:
                messages.add_message(request, messages.ERROR, 'Informacion incompleta')
                return HttpResponseRedirect(reverse('usuarios:generargasto'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class BuscarSolicitud(LoginRequiredMixin, View):
    login_url = '/'
    form_class = GastosForm
    template_name = 'usuarios/modificarestado.html'

    def get(self, request, IdSoGa):
        try:
            solicitud = SolicitudGastos.objects.get(IdSoGa=IdSoGa)
            lsg = SolicitudGastos.objects.filter(IdSoGa=IdSoGa)
            form = self.form_class(instance=solicitud)
            return render(request, self.template_name,
                          {
                              'lsg': lsg,
                              'form': form,
                              'estado': solicitud.Estado,
                              'IdSoGa': solicitud.IdSoGa
                          })

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request, IdSoGa):
        try:
            solicitud = SolicitudGastos.objects.get(IdSoGa=IdSoGa)
            form = self.form_class(request.POST, instance=solicitud)

            if form.is_valid():
                form.save()
                messages.add_message(request, messages.INFO, 'El estado de la orden de modifico correctamente')
                return HttpResponseRedirect(reverse('usuarios:controlpresupuestal'))

            else:
                messages.add_message(request, messages.ERROR,
                                     'No se puedo modificar el estado de la orden, verifique la informacion')
                return HttpResponseRedirect(reverse('usuarios:controlpresupuestal'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class ListasGastos(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/listasgastos.html'

    def get(self, request):
        try:
            usuario = Usuario.objects.get(usuid=request.user.pk)
            idacueducto = usuario.IdAcueducto
            contadorpen = SolicitudGastos.objects.filter(Estado=ESTADO1,IdAcueducto=idacueducto)
            solicitudesgastos = SolicitudGastos.objects.filter(IdAcueducto=idacueducto).order_by("-IdSoGa")
            cierre = Cierres.objects.filter(IdAcueducto=idacueducto)
            contador = SolicitudGastos.objects.filter(IdAcueducto=idacueducto).count()
            contcierre = Cierres.objects.filter(IdAcueducto=idacueducto).count()

            return render(request, self.template_name, {
                'solicitudesgastos': solicitudesgastos,
                'contador': contador,
                'cierres': cierre,
                'contcierre': contcierre,
                'notificaciones': contadorpen,
            })

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class RegistroMedidor(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/registromedidor.html'
    form_class = MedidoresForm

    def get(self, request):
        try:

            form = self.form_class()
            return render(request, self.template_name, {
                'form': form
            })

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request):
        form = self.form_class(request.POST)
        try:
            if form.is_valid():
                form.save()
                messages.add_message(request, messages.INFO, 'El medidor se asigno correctamente')
                return HttpResponseRedirect(reverse('usuarios:consumos'))

            else:
                messages.add_message(request, messages.ERROR, 'El predio ya tiene medidor asignado')
                return HttpResponseRedirect(reverse('usuarios:consumos'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class AsignarMedidor(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/asignacionmedidor.html'
    form_class = FormAsignarMedidor

    def get(self, request, IdMedidor):
        try:
            consulta1 = Medidores.objects.get(IdMedidor=IdMedidor)
            form = self.form_class(instance=consulta1)
            return render(request, self.template_name, {
                'form': form
            })

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request, IdMedidor):
        try:
            usuario = Usuario.objects.get(usuid=request.user.pk)
            idacueducto = usuario.IdAcueducto
            acueducto = Acueducto.objects.get(IdAcueducto=idacueducto)
            idmedidor = IdMedidor
            idvivienda = request.POST.get("IdVivienda")
            estado = request.POST.get("Estado")
            consultarpredio = Asignacion.objects.filter(IdVivienda=idvivienda, Estado='Operativo',IdAcueducto=idacueducto).exists()

            if consultarpredio is False:
                vivienda = Vivienda.objects.get(IdVivienda=idvivienda)
                medidor = Medidores.objects.get(IdMedidor=idmedidor)
                medidor.Estado ='Asignado'
                medidor.save()
                asignacion = Asignacion(IdMedidor=medidor, IdVivienda=vivienda, Estado='Operativo',IdAcueducto=acueducto)
                asignacion.save()
                vivienda.TipoRecaudo = 'Medicion'
                vivienda.save()
                messages.add_message(request, messages.INFO, 'El medidor se asigno correctamente')
                return HttpResponseRedirect(reverse('usuarios:consumos'))

            else:
                messages.add_message(request, messages.ERROR, 'El predio ya tiene un medidor asignado y operativo')
                return HttpResponseRedirect(reverse('usuarios:consumos'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class RegistroCostoM(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/registrocostomatricula.html'
    form_class = CostoMForm

    def get(self, request):
        try:
            form = self.form_class()

            usuario = Usuario.objects.get(usuid=request.user.pk)
            tipousuario = Permisos.objects.filter(usuid=usuario, TipoPermiso='AccessPanel').exists()
            if tipousuario is True:
                return render(request, self.template_name, {
                    'form': form
                })
            else:
                messages.add_message(request, messages.ERROR,
                                     'Su usuario no tiene los permiso de acceso a esta seccion')
                return HttpResponseRedirect(reverse('usuarios:paneladmin'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request):
        try:
            valor = request.POST.get("Valor")

            if valor is not None:
                poblacion = ValorMatricula(Valor=valor)
                poblacion.save()
                messages.add_message(request, messages.INFO, 'El valor se agrego correctamente')
                return HttpResponseRedirect(reverse('usuarios:paneladmin'))

            else:
                messages.add_message(request, messages.ERROR, 'Debe ingresar un valor al campo')
                return HttpResponseRedirect(reverse('usuarios:paneladmin'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class RegistroTarifa(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/registrotarifa.html'
    form_class = TarifasForm

    def get(self, request):
        try:
            usuario = Usuario.objects.get(usuid=request.user.pk)
            form = self.form_class()
            tipousuario = Permisos.objects.filter(usuid=usuario, TipoPermiso='AccessPanel').exists()
            if tipousuario is True:
                return render(request, self.template_name, {
                    'form': form
                })

            else:
                messages.add_message(request, messages.ERROR,'Su usuario no tiene los permiso de acceso a esta seccion')
                return HttpResponseRedirect(reverse('usuarios:paneladmin'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request):
        form = self.form_class(request.POST)
        try:
            if form.is_valid():
                form.save()
                messages.add_message(request, messages.INFO, 'La informacion se guardo correctamente')
                return HttpResponseRedirect(reverse('usuarios:paneladmin'))

            else:
                messages.add_message(request, messages.ERROR, 'Ya existe una tarifa con esa informacion')
                return HttpResponseRedirect(reverse('usuarios:paneladmin'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class ImprimirTiquet(LoginRequiredMixin):
    login_url = '/'

    def get(self, request):
        try:
            impresoras = ConectorV3.obtenerImpresoras()
            print("Las impresoras son:")
            print(impresoras)

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class RegistroPqr(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/registropqrs.html'
    form_class = FormRegistroPqrs

    def get(self, request):
        try:

            tipousuario = True
            if tipousuario is True:
                form = self.form_class()
                return render(request, self.template_name, {'form': form})
            else:
                messages.add_message(request, messages.ERROR,'Su usuario no tiene los permiso de acceso a esta seccion')
                return HttpResponseRedirect(reverse('usuarios:inicio'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request):
        try:
            usuario = Usuario.objects.get(usuid=request.user.pk)
            idacueducto = usuario.IdAcueducto
            acueducto = Acueducto.objects.get(IdAcueducto=idacueducto)
            nombre = request.POST.get("Nombre")
            celular = request.POST.get("Telefono")
            correo = request.POST.get("Correo")
            direccion = request.POST.get("Direccion")
            tiposolicitud = request.POST.get("TipoSolicitud")
            clasificacion = request.POST.get("Clasificacion")
            descripcion = request.POST.get("Descripcion")
            usuario = Usuario.objects.get(usuid=request.user.pk)
            pqr = Pqrs(Nombre=nombre, Telefono=celular, Descripcion=descripcion, Correo=correo, Direccion=direccion,
                       TipoSolicitud=tiposolicitud, Clasificacion=clasificacion, Estado='Pendiente', usuid=usuario, IdAcueducto=acueducto)
            pqr.save()
            idpqr = pqr.IdPqrs
            messages.add_message(request, messages.INFO,'La pqrs se ' + str(tiposolicitud) + ' registro correctamente, RADICADO No: ' + str(
                                     idpqr))
            return HttpResponseRedirect(reverse('usuarios:listapqrs'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class ListaPqrs(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/listapqrs.html'

    def get(self, request):
        try:
            usuario = Usuario.objects.get(usuid=request.user.pk)
            idacueducto = usuario.IdAcueducto
            acueducto = Acueducto.objects.get(IdAcueducto=idacueducto)
            listado = Pqrs.objects.filter(IdAcueducto=idacueducto).order_by("-IdPqrs")
            total = Pqrs.objects.filter(IdAcueducto=idacueducto).count()
            lista = Pqrs.objects.filter(Estado='Pendiente',IdAcueducto=idacueducto)
            contcerrada = Pqrs.objects.filter(Estado='Cerrada',IdAcueducto=idacueducto).count()
            contpendiente = Pqrs.objects.filter(Estado='Pendiente',IdAcueducto=idacueducto).count()
            # tipo de solicitud
            contpeticion = Pqrs.objects.filter(TipoSolicitud='Peticion',IdAcueducto=idacueducto).count()
            contquejas = Pqrs.objects.filter(TipoSolicitud='Queja',IdAcueducto=idacueducto).count()
            contsolicitud = Pqrs.objects.filter(TipoSolicitud='Solicitud',IdAcueducto=idacueducto).count()
            contreclamo = Pqrs.objects.filter(TipoSolicitud='Reclamo',IdAcueducto=idacueducto).count()

            tipousuario = True
            if tipousuario is True:
                return render(request, self.template_name, {
                    'lista': lista,
                    'total': total,
                    'contp': contpeticion,
                    'contq': contquejas,
                    'conts': contsolicitud,
                    'contr': contreclamo,
                    'contpen': contpendiente,
                    'contcerrada': contcerrada,
                    'listado': listado
                })
            else:
                messages.add_message(request, messages.ERROR,'Su usuario no tiene los permiso de acceso a esta seccion')
                return HttpResponseRedirect(reverse('usuarios:inicio'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class VerPqr(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/verpqr.html'

    def get(self, request, idpqr):
        try:
            pqr = Pqrs.objects.filter(IdPqrs=idpqr)
            respuesta = RespuestasPqrs.objects.filter(IdPqrs=idpqr)
            idsolicitud = Pqrs.objects.get(IdPqrs=idpqr)

            return render(request, self.template_name, {
                'pqr': pqr,
                'respuestas': respuesta,
                'idsolicitud': idsolicitud
            })

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class RespuestaPqrs(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/respuestapqr.html'
    form_class = FormRespuestaPqrs

    def get(self, request, idsolicitud):
        try:
            prq = Pqrs.objects.get(IdPqrs=idsolicitud)
            form = self.form_class()

            return render(request, self.template_name, {
                'form': form,
                'idsolicitud': idsolicitud
            })

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request, idsolicitud):
        try:
            usuario = Usuario.objects.get(usuid=request.user.pk)
            idacueducto = usuario.IdAcueducto
            acueducto = Acueducto.objects.get(IdAcueducto=idacueducto)
            id = 1
            soporte = request.FILES.get("Soporte")
            descripcion = request.POST.get("Descripcion")
            pqrs = Pqrs.objects.get(IdPqrs=idsolicitud)
            if id>=1:
                cpqrs = Pqrs.objects.get(IdPqrs=idsolicitud)
                cpqrs.Estado = NOVEDAD1
                cpqrs.save()
                respuesta = RespuestasPqrs(IdPqrs=pqrs, Descripcion=descripcion, Soporte=soporte, IdAcueducto=acueducto)
                respuesta.save()
                messages.add_message(request, messages.INFO, 'La respuesta se agrego correctamente')
                return HttpResponseRedirect(reverse('usuarios:listapqrs'))

            else:
                messages.add_message(request, messages.WARNING, 'No se pudo agregar la respuesta')
                return HttpResponseRedirect(reverse('usuarios:listapqrs'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

class Suspenciones(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/generadorsuspenciones.html'

    def get(self, request):
        try:
            usuario = Usuario.objects.get(usuid=request.user.pk)
            idacueducto = usuario.IdAcueducto
            cantanuladas = OrdenesTrabajo.objects.filter(Estado=SA, TipoNovedad='Suspension',IdAcueducto=idacueducto).count()
            cantejecutadas = OrdenesTrabajo.objects.filter(Estado=SJ, TipoNovedad='Suspension',IdAcueducto=idacueducto).count()
            cantpendientes = OrdenesTrabajo.objects.filter(Estado=SP, TipoNovedad='Suspension',IdAcueducto=idacueducto).count()
            ordenessuspenciones = OrdenesTrabajo.objects.filter(Estado=SP, TipoNovedad='Suspension',IdAcueducto=idacueducto)
            totales = OrdenesTrabajo.objects.filter(TipoNovedad='Suspension',IdAcueducto=idacueducto).count()

            tipousuario = True

            if tipousuario is True:
                return render(request, self.template_name, {
                    'anuladas': cantanuladas,
                    'pendientes': cantpendientes,
                    'ejecutadas': cantejecutadas,
                    'ordsus': ordenessuspenciones,
                    'total': totales

                })
            else:
                messages.add_message(request, messages.ERROR,
                                     'Su usuario no tiene los permiso de acceso a esta seccion')
                return HttpResponseRedirect(reverse('usuarios:inicio'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

class Reconexiones(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/generadorreconexiones.html'

    def get(self, request):
        try:
            usuario = Usuario.objects.get(usuid=request.user.pk)
            idacueducto = usuario.IdAcueducto
            total = OrdenesTrabajo.objects.filter(TipoNovedad='Reconexion',IdAcueducto=idacueducto).count()
            ordenesreconexion = OrdenesTrabajo.objects.filter(Estado=SP,TipoNovedad='Reconexion',IdAcueducto=idacueducto)
            contreeje = OrdenesTrabajo.objects.filter(Estado='Cerrada',TipoNovedad='Reconexion',IdAcueducto=idacueducto).count()
            contrepen = OrdenesTrabajo.objects.filter(Estado=SP,TipoNovedad='Reconexion',IdAcueducto=idacueducto).count()
            tipousuario = True

            if tipousuario is True:
                return render(request, self.template_name, {
                    'ordrec': ordenesreconexion,
                    'rependientes': contrepen,
                    'reejecutadas': contreeje,
                    'total': total

                })
            else:
                messages.add_message(request, messages.ERROR,
                                     'Su usuario no tiene los permisos de acceso a esta seccion')
                return HttpResponseRedirect(reverse('usuarios:inicio'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class VerOrdenSuspencion(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/verordensuspencion.html'

    def get(self, request, IdOrden):
        try:
            ordenessuspencion = OrdenesTrabajo.objects.get(IdOrden=IdOrden)
            ots = OrdenesTrabajo.objects.filter(IdOrden=IdOrden)
            idorden = ordenessuspencion.IdOrden
            estado = ordenessuspencion.Estado

            return render(request, self.template_name,
                          {'idorden': idorden,
                           'estado': estado,
                           'ots': ots,
                           })

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request, IdOrden):
        try:
            usuario = Usuario.objects.get(usuid=request.user.pk)
            usu = str(usuario)
            idorden = IdOrden
            idacueducto = usuario.IdAcueducto
            acueducto = Acueducto.objects.get(IdAcueducto=idacueducto)
            idtarifa = acueducto.IdTarifa
            tarifa = Tarifa.objects.get(IdTarifa=idtarifa.pk)
            valorsuspencion = tarifa.TarifaSuspencion
            orden = OrdenesTrabajo.objects.filter(IdOrden=IdOrden).exists()
            otra = OrdenesTrabajo.objects.get(IdOrden=IdOrden)
            descripcion = 'Suspensión'
            s = (datetime.today())
            if orden is True:
                ordensuspencion = OrdenesTrabajo.objects.get(IdOrden=IdOrden)
                ordensuspencion.Estado = SJ
                ordensuspencion.FechaEjecucion = s
                ordensuspencion.usuario = usu
                ordensuspencion.save()
                idvivienda = ordensuspencion.IdVivienda
                vivienda = Vivienda.objects.get(IdVivienda=idvivienda.pk)
                vivienda.EstadoServicio = E2
                vivienda.save()
                concepto = Conceptos(Tipo=descripcion, Observacion='OTS: '+idorden, Estado='Sin facturar',
                                     Valor=valorsuspencion, IdVivienda=idvivienda)
                concepto.save()
                messages.add_message(request, messages.INFO, 'La orden se cerro correctamente')
                return HttpResponseRedirect(reverse('usuarios:suspenciones'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class AnularFactura(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/anularfactura.html'

    def get(self, request):
        try:
            usuario = Usuario.objects.get(usuid=request.user.pk)
            listapqrs = Pqrs.objects.filter(Estado='Pendiente')
            contqrs = Pqrs.objects.filter(Estado='Pendiente').count()
            contsoli = SolicitudGastos.objects.filter(Estado=ESTADO1).count()
            totalnoti = contqrs + contsoli
            contadorpen = SolicitudGastos.objects.filter(Estado=ESTADO1)
            tipousuario = Permisos.objects.filter(usuid=usuario, TipoPermiso='AF').exists()
            if tipousuario is True:
                return render(request, self.template_name, {
                    'notificaciones': contadorpen,
                    'listapqrs': listapqrs,
                    'totalnoti': totalnoti
                })
            else:
                messages.add_message(request, messages.ERROR,
                                     'Su usuario no tiene los permisos de acceso a esta seccion')
                return HttpResponseRedirect(reverse('usuarios:inicio'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request):
        try:
            numerofactura = request.POST.get("factura", "")
            factura = Facturas.objects.filter(IdFactura=numerofactura).exists()
            if factura is True:
                fac = Facturas.objects.get(IdFactura=numerofactura)
                fac.Estado = FA
                fac.save()
                messages.add_message(request, messages.INFO, 'La factura se anulo correctamente')
                return HttpResponseRedirect(reverse('usuarios:anularfactura'))

            else:
                messages.add_message(request, messages.WARNING, 'El numero de factura ingresado no existe')
                return HttpResponseRedirect(reverse('usuarios:anularfactura'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class DescargarFactura(LoginRequiredMixin, View):
    login_url = '/'

    def get(self, request, IdFactura):
        try:
            # datos factura
            conceptos1 = FacturasConceptos.objects.get(IdFactura=IdFactura)
            idconcepto = conceptos1.IdConcepto
            concepto = ConceptosFacturados.objects.get(IdRegistro=idconcepto.pk)
            aporte = concepto.AporteFijo
            basico = concepto.Basico
            complementario = concepto.Complementario
            suspencion = concepto.Suspencion
            reconexion = concepto.Reconexion
            recargo = concepto.Recargo
            acuerdopago = concepto.AcuerdoPago
            saldoanterior = concepto.SaldoAnterior
            subsidio = concepto.Subsidio
            # totales acueducto
            sumaacueducto = aporte + basico + complementario + suspencion + reconexion + recargo - subsidio

            # datos factura
            factura = Facturas.objects.get(IdFactura=IdFactura)
            noaporte = factura.IdFactura
            estado = factura.Estado
            vencidas = factura.facturasvencidas
            mes = factura.periodofacturado
            periodofacturado = factura.periodofacturado
            FechaExpe = factura.FechaExpe.date()
            FechaLimite = factura.FechaLimite.date()
            Total = factura.Total
            matricula = factura.IdVivienda

            # identificador de vivienda
            vivienda = Vivienda.objects.get(IdVivienda=matricula.pk)
            idmatricula = vivienda.IdVivienda
            idtitular = vivienda.IdPropietario
            sector = vivienda.Direccion
            casa = vivienda.NumeroCasa
            piso = vivienda.Piso
            direccion = sector + ' Cs ' + casa + ' Ps ' + piso
            estrato = vivienda.Estrato
            tipoinstalacion = vivienda.TipoInstalacion
            tipodepredio = vivienda.InfoInstalacion
            estadoservicio = vivienda.EstadoServicio
            diametro = vivienda.Diametro

            # identificador de propietario
            titular = Propietario.objects.get(IdPropietario=idtitular.pk)
            nombretitular = titular.Nombres
            apellidotitular = titular.Apellidos
            nombrecompleto = nombretitular + ' ' + apellidotitular

            # libro excel
            condicion=0
            if condicion == 0:
                qr = qrcode.QRCode(
                    version=6,
                    error_correction=qrcode.constants.ERROR_CORRECT_L,
                    box_size=3,
                    border=0,
                )
                qr.add_data(noaporte)
                qr.make(fit=True)
                imga = qr.make_image(fill_color="black", back_color="white")
                imga.save('static/ModeloFactura/output.png')
                wb = openpyxl.load_workbook('static/ModeloFactura/003-0-240824v11.xlsx')
                ws = wb.active
                img = openpyxl.drawing.image.Image('static/ModeloFactura/output.png')
                ws.add_image(img, 'AA9')
                if int(saldoanterior) >= 1:
                    imagen = openpyxl.drawing.image.Image('static/ModeloFactura/corte1.png')
                    ws.add_image(imagen, 'AJ16')
                else:
                    pass
                # factura matricula estado
                ws['AZ1'] = noaporte
                ws['A15'] = idmatricula
                ws['AZ3'] = estado
                # suscriptor
                ws['H15'] = nombrecompleto
                ws['A13'] = direccion
                ws['A11'] = direccion
                ws['P13'] = diametro
                ws['P11'] = estrato
                ws['T13'] = tipoinstalacion
                ws['T11'] = tipodepredio
                ws['S9'] = estadoservicio
                # Periodo facturado
                ws['S6'] = periodofacturado

                # ultimo pago
                consultarpago = Pagos.objects.filter(IdVivienda=matricula).exists()
                if consultarpago is True:
                    filtropagos = Pagos.objects.filter(IdVivienda=matricula).order_by("-IdPago")[:1]
                    consultarp = Pagos.objects.get(IdPago=filtropagos)
                    ws['H18'] = str(consultarp.IdPago) + " - " + str(consultarp.FechaPago.date()) + " - $"+str(consultarp.ValorPago)
                else:
                    mensaje = "No Registra"
                    ws['H18'] = mensaje

                # conceptos facturados
                if int(saldoanterior) > 0:
                    ws['BB7'] = int(saldoanterior)

                if int(aporte) > 0:
                    ws['BB8'] = int(aporte)

                if int(basico) > 0:
                    consumo = Consumos.objects.get(IdVivienda=matricula,mes=mes)
                    if consumo.Consumo <=20 and consumo.Consumo >0:
                        ws['AQ9'] = consumo.Consumo
                        ws['BB9'] = int(basico)
                    elif consumo.Consumo >=21:
                        ws['AQ9'] = 20
                        ws['BB9'] = int(basico)

                if int(complementario) > 0:
                    consumo = Consumos.objects.get(IdVivienda=matricula,mes=mes)
                    ws['AQ10'] = consumo.Consumo - 20
                    ws['BB10'] = int(complementario)

                if int(suspencion) > 0:
                    ws['BB11'] = int(suspencion)

                if int(reconexion) > 0:
                    ws['BB12'] = int(reconexion)

                if int(recargo) > 0:
                    ws['BB13'] = int(recargo)

                if int(acuerdopago) > 0:
                    acuerdo1 = AcuerdosPago.objects.filter(IdVivienda=matricula, Estado='Pendiente').exists()
                    if acuerdo1 is True:
                        acuerdo = AcuerdosPago.objects.get(IdVivienda=matricula, Estado='Pendiente')
                        ws['AQ19'] = int(acuerdo.ValorPendiente)
                        ws['AW19'] = acuerdo.CuotasPendientes
                        ws['BA19'] = int(acuerdopago)
                        ws['BB20'] = int(acuerdopago)
                    else:
                        ws['AQ19'] = 0
                        ws['AW19'] = 1
                        ws['BA19'] = int(acuerdopago)
                        ws['BB20'] = int(acuerdopago)

                if int(subsidio) > 0:
                    ws['BB14'] = int(subsidio)

                # total concepto de acueducto
                ws['BB15'] = int(sumaacueducto) + int(saldoanterior)

                # facturas vencidas
                ws['AB24'] = vencidas

                # fechas de procedimiento
                ws['AK30'] = FechaExpe
                ws['AK31'] = FechaLimite

                if int(vencidas) >= 1:
                    ws['AK31'] = 'Inmediato'
                    ws['AK32'] = FechaLimite

                # total a pagar condional 0
                if int(Total) <= 0:
                    ws['AS32'] = 0
                else:
                    ws['AS32'] = int(Total)

                medidor = Asignacion.objects.filter(IdVivienda=matricula).exists()
                if medidor is True:
                    consumo = Consumos.objects.get(IdVivienda=matricula,mes=mes)
                    ws['H21'] = str(consumo.IdMedidor)
                    ws['H22'] = consumo.Lecturaactual
                    ws['H23'] = consumo.Lecturaanterior
                    ws['H24'] = consumo.Consumo
                    ws['H25'] = consumo.promedio
                    ws['H26'] = consumo.diasconsumo
                    ws['S21'] = consumo.mes
                    cantidad = consumo.Consumo
                    if cantidad == 0:
                        imagen2 = openpyxl.drawing.image.Image('static/images/emoji4.png')
                        ws.add_image(imagen2, 'P22')

                    elif cantidad >=11 and cantidad <=20:
                        imagen2 = openpyxl.drawing.image.Image('static/images/emoji4.png')
                        ws.add_image(imagen2, 'P22')

                    elif cantidad >=21:
                        imagen2 = openpyxl.drawing.image.Image('static/images/emoji.png')
                        ws.add_image(imagen2, 'P22')

                ws.title = IdFactura
                archivo_predios = str(IdFactura) + ".xlsx"
                response = HttpResponse(content_type="application/ms-excel")
                content = "attachment; filename = {0}".format(archivo_predios)
                response['Content-Disposition'] = content
                wb.save(response)
                return response

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class VerOrdenReconexion(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/verordenreconexion.html'

    def get(self, request, IdOrden):
        try:
            ordenesreconexicion = OrdenesTrabajo.objects.get(IdOrden=IdOrden)
            otr = OrdenesTrabajo.objects.filter(IdOrden=IdOrden)
            idorden = ordenesreconexicion.IdOrden
            estado = ordenesreconexicion.Estado

            return render(request, self.template_name,
                          {'idorden': idorden,
                           'estado': estado,
                           'otr': otr
                           })

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request, IdOrden):
        try:
            usuario = Usuario.objects.get(usuid=request.user.pk)
            usu = str(usuario)
            idorden = IdOrden
            idacueducto = usuario.IdAcueducto
            acueducto = Acueducto.objects.get(IdAcueducto=idacueducto)
            idtarifa = acueducto.IdTarifa
            tarifa = Tarifa.objects.get(IdTarifa=idtarifa)
            valorreconexion = tarifa.TarifaReconexion
            estado = ESTADO1
            descripcion = 'Reconexion'
            otra = OrdenesTrabajo.objects.get(IdOrden=IdOrden)
            idestadocuenta = otra.IdEstadoCuenta
            orden = OrdenesTrabajo.objects.filter(IdOrden=IdOrden).exists()
            s = (datetime.today())
            fecha = s
            if orden is True:
                ordensuspencion = OrdenesTrabajo.objects.get(IdOrden=IdOrden)
                ordensuspencion.Estado = 'Cerrada'
                ordensuspencion.FechaEjecucion = fecha
                ordensuspencion.usuario = usu
                ordensuspencion.save()
                idvivienda = ordensuspencion.IdVivienda
                vivienda = Vivienda.objects.get(IdVivienda=idvivienda)
                vivienda.EstadoServicio = E1
                vivienda.save()
                concepto = Conceptos(Tipo=descripcion, Observacion='OTS: ' + idorden, Estado='Sin facturar',
                                     Valor=valorreconexion, IdVivienda=idvivienda.pk)
                concepto.save()
                messages.add_message(request, messages.INFO,
                                     'La orden se cerro correctamente y el predio '
                                     'cambio de estado suspendido a operativo')
                return HttpResponseRedirect(reverse('usuarios:suspenciones'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class GeneradorFacturasIndividual(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/facturaindividual.html'

    def get(self, request, IdVivienda):
        try:
            estadoscuenta = EstadoCuenta.objects.filter(IdVivienda=IdVivienda)
            matricula = IdVivienda
            verificacion = CobroMatricula.objects.filter(IdVivienda=matricula, Estado=ESTCOBRO)
            return render(request, self.template_name,
                          {
                              'estadocuenta': estadoscuenta,
                              'matricula': matricula,
                              'matriculas': verificacion
                          })

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request, IdVivienda):
        try:
            fechaexp = (datetime.today())
            ciclo = fechaexp.month
            ciclos = Ciclo.objects.get(IdCiclo=ciclo)
            fechalimite = fechaexp + timedelta(days=DIASFACTURACION)
            estadoc = EstadoCuenta.objects.get(IdVivienda=IdVivienda)
            total = estadoc.Valor
            facturas = Facturas.objects.filter(IdEstadoCuenta=estadoc, Estado=EF).count()
            if facturas >= 1:
                messages.add_message(request, messages.WARNING, 'Ya existe una factura pendiente de pago')
                return HttpResponseRedirect(reverse('usuarios:inicio'))

            else:
                verificacion = CobroMatricula.objects.get(IdVivienda=IdVivienda)
                if verificacion.Estado == ESTCOBRO:
                    otrocosto = verificacion.Cuota
                    final = int(total) + int(otrocosto)
                    factura = Facturas(Estado=EF, FechaExpe=fechaexp, FechaLimite=fechalimite, Total=final,
                                      IdCiclo=ciclos,
                                      IdEstadoCuenta=estadoc, TotalConsumo=total, OtrosCobros=otrocosto)
                    factura.save()
                    messages.add_message(request, messages.INFO, 'La factura se creo correctamente')
                    return HttpResponseRedirect(reverse('usuarios:inicio'))

                else:
                    factura = Facturas(Estado=EF, FechaExpe=fechaexp, FechaLimite=fechalimite,
                                      Total=total, IdCiclo=ciclos, IdEstadoCuenta=estadoc, TotalConsumo=total,
                                      OtrosCobros=0)
                    factura.save()
                    messages.add_message(request, messages.INFO, 'La factura se creo correctamente')
                    return HttpResponseRedirect(reverse('usuarios:inicio'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class DescargaMasivaFacturas(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/listamasivafacturas.html'

    def get(self, request):
        try:
            usuario = Usuario.objects.get(usuid=request.user.pk)
            idacueducto = usuario.IdAcueducto
            total = Facturas.objects.filter(Estado='Emitida', IdAcueducto=idacueducto).count()
            facturas = Facturas.objects.filter(Estado=EF, IdAcueducto=idacueducto).order_by('IdFactura')
            return render(request, self.template_name, {
                'facturas': facturas,
                'total': total,
            })

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class ReportePdfPagos(LoginRequiredMixin, View):
    login_url = '/'

    def get(self, request):
        usuario = Usuario.objects.get(usuid=request.user.pk)
        idacueducto = usuario.IdAcueducto
        pagos = Pagos.objects.filter(IdAcueducto=idacueducto)
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de pagos"
        ws['A1'] = 'Referencia de pago'
        ws['B1'] = 'Numero de factura'
        ws['C1'] = 'Año'
        ws['D1'] = 'Valor pagado'
        ws['E1'] = 'Efectivo'
        ws['F1'] = 'Cambio'
        ws['G1'] = 'Usuario recaudo'
        ws['H1'] = 'Matricula'
        ws['I1'] = 'Fecha de pago'
        sfecha = (datetime.today())
        cont = 2

        for pago in pagos:
            ws.cell(row=cont, column=1).value = pago.IdPago
            ws.cell(row=cont, column=2).value = str(pago.IdFactura)
            ws.cell(row=cont, column=3).value = pago.Ano
            ws.cell(row=cont, column=4).value = pago.ValorPago
            ws.cell(row=cont, column=5).value = pago.Efectivo
            ws.cell(row=cont, column=6).value = pago.Devuelta
            ws.cell(row=cont, column=7).value = str(pago.IdUsuario)
            ws.cell(row=cont, column=8).value = str(pago.IdVivienda)
            ws.cell(row=cont, column=9).value = str(pago.FechaPago)

            cont += 1

        archivo_propi = "ReportePagos" + str(sfecha) + ".xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(archivo_propi)
        response['Content-Disposition'] = content
        wb.save(response)
        return response

class ReporteConceptos(LoginRequiredMixin, View):
    login_url = '/'

    def get(self, request):
        usuario = Usuario.objects.get(usuid=request.user.pk)
        idacueducto = usuario.IdAcueducto
        pagos = ConceptosFacturados.objects.filter(IdAcueducto=idacueducto)
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de pagos"
        ws['A1'] = 'No registro'
        ws['B1'] = 'Aporte fijo'
        ws['C1'] = 'Cuota matricula'
        ws['D1'] = 'Suspension'
        ws['E1'] = 'Reconexion'
        ws['F1'] = 'Acuerdo Pago'
        ws['G1'] = 'Subsidio'
        ws['H1'] = 'Fecha'
        ws['I1'] = 'Total'
        ws['J1'] = 'Estado'
        ws['K1'] = 'Periodo'
        ws['L1'] = 'Matricula'
        ws['M1'] = 'Complementario'
        ws['N1'] = 'Recargo'
        ws['O1'] = 'Saldo anterior'

        sfecha = (datetime.today())
        cont = 2

        for pago in pagos:
            ws.cell(row=cont, column=1).value = pago.IdRegistro
            ws.cell(row=cont, column=2).value = str(pago.AporteFijo)
            ws.cell(row=cont, column=3).value = str(pago.CuotaMatricula)
            ws.cell(row=cont, column=4).value = str(pago.Suspencion)
            ws.cell(row=cont, column=5).value = str(pago.Reconexion)
            ws.cell(row=cont, column=6).value = str(pago.AcuerdoPago)
            ws.cell(row=cont, column=7).value = str(pago.Subsidio)
            ws.cell(row=cont, column=8).value = str(pago.Fecha)
            ws.cell(row=cont, column=9).value = str(pago.Total)
            ws.cell(row=cont, column=10).value = str(pago.Estado)
            ws.cell(row=cont, column=11).value = str(pago.Periodo)
            ws.cell(row=cont, column=12).value = str(pago.IdVivienda)
            ws.cell(row=cont, column=13).value = str(pago.Complementario)
            ws.cell(row=cont, column=14).value = str(pago.Recargo)
            ws.cell(row=cont, column=15).value = str(pago.SaldoAnterior)
            cont += 1

        archivo_propi = "ReporteConceptos" + str(sfecha) + ".xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(archivo_propi)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


class ReporteCompleto(LoginRequiredMixin, View):
    login_url = '/'

    def get(self, request):
        usuario = Usuario.objects.get(usuid=request.user.pk)
        idacueducto = usuario.IdAcueducto
        viviendas = Vivienda.objects.filter(IdAcueducto=idacueducto)
        sfecha = (datetime.today())
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de predios"
        ws['A1'] = 'Matricula'
        ws['B1'] = 'Dirección'
        ws['C1'] = 'NumeroCasa'
        ws['D1'] = 'Piso'
        ws['E1'] = 'Ciclo'
        ws['F1'] = 'TipoInstalación'
        ws['G1'] = 'Estrato'
        ws['H1'] = 'EstadoServicio'
        ws['I1'] = 'Acueducto'
        ws['J1'] = 'Usuario'
        ws['K1'] = 'MatriculaAnt'
        ws['L1'] = 'InfoInstalación'
        ws['M1'] = 'ProfAcometida'
        ws['N1'] = 'CantHabitantes'
        ws['O1'] = 'Numero de cedula'
        ws['P1'] = 'Nombres'
        ws['Q1'] = 'Apellidos'
        ws['R1'] = 'Telefono'
        ws['S1'] = 'Email'
        ws['T1'] = 'Tipo de poblacion'
        ws['U1'] = 'Ficha catastral'

        cont = 2

        for vivienda in viviendas:
            idpropietario = vivienda.IdPropietario
            ws.cell(row=cont, column=1).value = vivienda.IdVivienda
            ws.cell(row=cont, column=2).value = vivienda.Direccion
            ws.cell(row=cont, column=3).value = vivienda.NumeroCasa
            ws.cell(row=cont, column=4).value = vivienda.Piso
            ws.cell(row=cont, column=5).value = vivienda.Ciclo
            ws.cell(row=cont, column=6).value = vivienda.TipoInstalacion
            ws.cell(row=cont, column=7).value = vivienda.Estrato
            ws.cell(row=cont, column=8).value = vivienda.EstadoServicio
            ws.cell(row=cont, column=9).value = str(vivienda.IdAcueducto)
            ws.cell(row=cont, column=10).value = str(vivienda.usuid)
            ws.cell(row=cont, column=11).value = vivienda.MatriculaAnt
            ws.cell(row=cont, column=12).value = vivienda.InfoInstalacion
            ws.cell(row=cont, column=13).value = vivienda.ProfAcometida
            ws.cell(row=cont, column=14).value = vivienda.CantHabitantes
            propietario = Propietario.objects.get(IdPropietario=idpropietario.pk)
            ws.cell(row=cont, column=15).value = propietario.IdPropietario
            ws.cell(row=cont, column=16).value = propietario.Nombres
            ws.cell(row=cont, column=17).value = propietario.Apellidos
            ws.cell(row=cont, column=18).value = propietario.NoTelefono
            ws.cell(row=cont, column=19).value = propietario.Email
            ws.cell(row=cont, column=20).value = str(propietario.IdPoblacion)
            ws.cell(row=cont, column=21).value = vivienda.FichaCastral
            cont += 1

        archivo_predios = "ReporteCompleto" + str(sfecha) + ".xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(archivo_predios)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


class ReporteSuspenciones(LoginRequiredMixin, View):
    login_url = '/'

    def get(self, request):
        usuario = Usuario.objects.get(usuid=request.user.pk)
        idacueducto = usuario.IdAcueducto
        pagos = OrdenesTrabajo.objects.filter(Estado=SP, TipoNovedad='Suspension', IdAcueducto=idacueducto)
        sfecha = (datetime.today())
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"
        ws['A1'] = 'Id Orden'
        ws['B1'] = 'Deuda'
        ws['C1'] = 'Fecha expedicion'
        ws['D1'] = 'Fecha ejecucion'
        ws['E1'] = 'Estado'
        ws['F1'] = 'Usuario encargado'
        ws['G1'] = 'Matricula'
        ws['H1'] = 'Sector'
        ws['I1'] = 'Casa'
        ws['J1'] = 'Piso'
        ws['K1'] = 'Estado servicio'
        ws['L1'] = 'Titular'
        cont = 2
        for suspencion in pagos:
            ws.cell(row=cont, column=1).value = suspencion.IdOrden
            ws.cell(row=cont, column=2).value = suspencion.Deuda
            ws.cell(row=cont, column=3).value = suspencion.FechaExpe
            ws.cell(row=cont, column=4).value = suspencion.FechaEjecucion
            ws.cell(row=cont, column=5).value = suspencion.Estado
            ws.cell(row=cont, column=6).value = suspencion.usuario
            idvivienda = suspencion.IdVivienda
            vivienda = Vivienda.objects.get(IdVivienda=idvivienda.pk)
            ws.cell(row=cont, column=7).value = vivienda.IdVivienda
            ws.cell(row=cont, column=8).value = vivienda.Direccion
            ws.cell(row=cont, column=9).value = vivienda.NumeroCasa
            ws.cell(row=cont, column=10).value = vivienda.Piso
            ws.cell(row=cont, column=11).value = vivienda.EstadoServicio
            ws.cell(row=cont, column=12).value = str(vivienda.IdPropietario)
            cont += 1

        archivo_predios = "Reporte ordenes suspencion" + str(sfecha) + ".xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(archivo_predios)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


class ReporteReconexion(LoginRequiredMixin, View):
    login_url = '/'

    def get(self, request):
        usuario = Usuario.objects.get(usuid=request.user.pk)
        idacueducto = usuario.IdAcueducto
        pagos = OrdenesTrabajo.objects.filter(Estado=SP, TipoNovedad='Reconexion', IdAcueducto=idacueducto)
        sfecha = (datetime.today())
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"
        ws['A1'] = 'Id Orden'
        ws['B1'] = 'Deuda'
        ws['C1'] = 'Fecha expedicion'
        ws['D1'] = 'Fecha ejecucion'
        ws['E1'] = 'Generado'
        ws['F1'] = 'Estado'
        ws['G1'] = 'Usuario encargado'
        ws['H1'] = 'Referencia'
        ws['I1'] = 'Matricula'
        ws['J1'] = 'Sector'
        ws['K1'] = 'Casa'
        ws['L1'] = 'Piso'
        ws['M1'] = 'Ciclo'
        ws['N1'] = 'Tipo Instalacion'
        ws['O1'] = 'Estrato'
        ws['P1'] = 'Estado servicio'
        ws['Q1'] = 'Titular'
        ws['R1'] = 'Info instalacion'
        ws['S1'] = 'Profundidad acometida'
        ws['T1'] = 'Cant habitantes'
        cont = 2
        for suspencion in pagos:
            ws.cell(row=cont, column=1).value = suspencion.IdOrden
            ws.cell(row=cont, column=2).value = suspencion.Deuda
            ws.cell(row=cont, column=3).value = suspencion.FechaExpe
            ws.cell(row=cont, column=4).value = suspencion.FechaEjecucion
            ws.cell(row=cont, column=5).value = suspencion.Generado
            ws.cell(row=cont, column=6).value = suspencion.Estado
            ws.cell(row=cont, column=7).value = suspencion.UsuarioEjecuta
            ws.cell(row=cont, column=8).value = str(suspencion.IdEstadoCuenta)
            idestadocuenta = suspencion.IdEstadoCuenta
            estadocuenta = EstadoCuenta.objects.get(IdEstadoCuenta=idestadocuenta.pk)
            idvivienda = estadocuenta.IdVivienda
            vivienda = Vivienda.objects.get(IdVivienda=idvivienda.pk)
            ws.cell(row=cont, column=9).value = vivienda.IdVivienda
            ws.cell(row=cont, column=10).value = vivienda.Direccion
            ws.cell(row=cont, column=11).value = vivienda.NumeroCasa
            ws.cell(row=cont, column=12).value = vivienda.Piso
            ws.cell(row=cont, column=13).value = vivienda.Ciclo
            ws.cell(row=cont, column=14).value = vivienda.TipoInstalacion
            ws.cell(row=cont, column=15).value = vivienda.Estrato
            ws.cell(row=cont, column=16).value = vivienda.EstadoServicio
            ws.cell(row=cont, column=17).value = str(vivienda.IdPropietario)
            ws.cell(row=cont, column=18).value = vivienda.InfoInstalacion
            ws.cell(row=cont, column=19).value = vivienda.ProfAcometida
            ws.cell(row=cont, column=20).value = vivienda.CantHabitantes
            cont += 1

        archivo_predios = "Reporte ordenes reconexion" + str(sfecha) + ".xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(archivo_predios)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


class EnvioCorreo(LoginRequiredMixin):
    login_url = '/'

    def get(self, asunto, informacion, tipousuario):
        now = datetime.now()
        context = {'asunto': asunto, 'fecha': now, 'informacion': informacion, 'usuario': tipousuario}
        template = get_template('usuarios/correo.html')
        content = template.render(context)

        email = EmailMultiAlternatives(
            asunto,
            'Sistemas acueducto caimalito',
            settings.EMAIL_HOST_USER,
            [username]
        )
        email.attach_alternative(content, 'text/html')
        email.send()

class CambioTitular(LoginRequiredMixin, View):
    login_url = '/'
    vervivienda = VisualizarVivienda
    template_name = 'usuarios/cambiotitular.html'
    predio = VisualizarVivienda

    def get(self, request, IdVivienda):
        try:
            vivienda = Vivienda.objects.get(IdVivienda=IdVivienda)
            idpropietario = vivienda.IdPropietario
            matricula = IdVivienda
            sector = vivienda.Direccion
            casa = vivienda.NumeroCasa
            infovivienda = matricula + " " + sector + " " + casa
            propietario = Propietario.objects.filter(IdPropietario=idpropietario.pk)
            usuario = Usuario.objects.get(usuid=request.user.pk)

            tipousuario = Permisos.objects.filter(usuid=usuario, TipoPermiso='NOACFA').exists()
            if tipousuario is False:
                return render(request, self.template_name, {'propietario': propietario, 'matricula': matricula,
                                                            'informacion': infovivienda})
        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request, IdVivienda):
        try:
            idpropietario = request.POST.get("idpropietario")
            propietario = Propietario.objects.filter(IdPropietario=idpropietario).exists()
            vivienda = Vivienda.objects.get(IdVivienda=IdVivienda)
            idpropi = vivienda.IdPropietario.pk
            if idpropi == idpropietario:
                ver = self.predio()
                messages.add_message(request, messages.ERROR,
                                     'el documento del titular ingresado es el mismo que tiene asignado '
                                     'actualmente el predio, modifique la informacion del titular directamente')
                ejecutar = ver.get(request, IdVivienda)
                return ejecutar

            else:
                if propietario is True:
                    usuario = Usuario.objects.get(usuid=request.user.pk)
                    idacueducto = usuario.IdAcueducto
                    datosacueducto = Acueducto.objects.get(IdAcueducto=idacueducto)
                    vivienda = Vivienda.objects.get(IdVivienda=IdVivienda)
                    propie = Propietario.objects.get(IdPropietario=idpropietario)
                    propie2 = Propietario.objects.get(IdPropietario=idpropi)
                    vivienda.IdPropietario = propie
                    vivienda.save()
                    tiponovedad = 'Cambio titular'
                    descripcion = 'se cambia titular de la matricula ' + str(IdVivienda) + ' de ' + str(
                        propie2) + ' por ' + str(propie) + ' por solicitud del interesado '
                    novedad = Novedades(Descripcion=descripcion, TipoNovedad=tiponovedad, usuario=usuario,
                                                 matricula=vivienda, IdAcueducto=datosacueducto)
                    novedad.save()
                    ver = self.predio()
                    messages.add_message(request, messages.INFO, 'Se modifica el titular correctamente')
                    ejecutar = ver.get(request, IdVivienda)
                    return ejecutar

                else:
                    ver = self.predio()
                    messages.add_message(request, messages.ERROR, 'el documento ingresado no existe en el sistema')
                    ejecutar = ver.get(request, IdVivienda)
                    return ejecutar

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class Mapa(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/mapa.html'

    def get(self, request):
        try:
            usuario = Usuario.objects.get(usuid=request.user.pk)
            tipousuario = Permisos.objects.filter(usuid=usuario, TipoPermiso='NOACMAPA').exists()
            drilistapqrs = Pqrs.objects.filter(Estado='Pendiente')
            contqrs = Pqrs.objects.filter(Estado='Pendiente').count()
            contsoli = SolicitudGastos.objects.filter(Estado=ESTADO1).count()
            totalnoti = contqrs + contsoli
            contadorpen = SolicitudGastos.objects.filter(Estado=ESTADO1)

            if tipousuario is False:
                return render(request, self.template_name, {
                    'notificaciones': contadorpen,
                    'listapqrs': drilistapqrs,
                    'totalnoti': totalnoti
                })

            else:
                messages.add_message(request, messages.ERROR,
                                     'Su usuario no tiene los permiso de acceso a esta seccion')
                return HttpResponseRedirect(reverse('usuarios:inicio'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class ModificarAcueducto(LoginRequiredMixin, View):
    login_url = '/'
    form_class = AcueductoAForm
    template_name = 'usuarios/modificarempresa.html'

    def get(self, request):
        try:
            usuario = Usuario.objects.get(usuid=request.user.pk)
            idacueducto = usuario.IdAcueducto
            datosacueducto = Acueducto.objects.get(IdAcueducto=idacueducto)
            form = self.form_class(instance=datosacueducto)


            tipousuario = Permisos.objects.filter(usuid=usuario, TipoPermiso='AccessPanel').exists()
            if tipousuario is True:
                return render(request, self.template_name,
                              {'form': form})

            else:
                messages.add_message(request, messages.ERROR,
                                     'Su usuario no tiene los permiso de acceso a esta seccion')
                return HttpResponseRedirect(reverse('usuarios:inicio'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request):
        try:
            usuario = Usuario.objects.get(usuid=request.user.pk)
            idacueducto = usuario.IdAcueducto
            datosacueducto = Acueducto.objects.get(IdAcueducto=idacueducto)
            form = self.form_class(request.user, request.POST, instance=datosacueducto)

            if form.is_valid():
                form.save()
                messages.add_message(request, messages.INFO, 'La información se modifico correctamente')
                return HttpResponseRedirect(reverse('usuarios:paneladmin'))

            else:
                messages.add_message(request, messages.ERROR, 'No se puedo modificar la informacion')
                return HttpResponseRedirect(reverse('usuarios:paneladmin'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

class DesactivarUsuarios(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/eliminarusuario.html'

    def get(self, request, usuid):
        try:
            usuario2 = Usuario.objects.get(usuid=request.user.pk)
            usuario = Usuario.objects.filter(usuid=usuid)
            user = User.objects.get(id=usuid)
            estado = user.is_active
            tipousuario = Permisos.objects.filter(usuid=usuario2, TipoPermiso='AccessPanel').exists()
            if tipousuario is True:
                return render(request, self.template_name, {'usuario': usuario, 'estado': estado})

            else:
                messages.add_message(request, messages.ERROR,
                                     'Su usuario no tiene los permiso de acceso a esta seccion')
                return HttpResponseRedirect(reverse('usuarios:inicio'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request, usuid):
        try:
            user = User.objects.get(id=usuid)
            estado = user.is_active
            if estado is True:
                user.is_active = False
                user.save()
                messages.add_message(request, messages.INFO, 'El usuario se desactivo correctamente')
                return HttpResponseRedirect(reverse('usuarios:paneladmin'))

            else:
                user.is_active = True
                user.save()
                messages.add_message(request, messages.INFO, 'El usuario se activo correctamente')
                return HttpResponseRedirect(reverse('usuarios:paneladmin'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class EliminarPoblacion(LoginRequiredMixin, View):
    login_url = '/'

    def get(self, request):
        try:
            identificador = request.GET.get("identificador")
            usuario = Usuario.objects.get(usuid=request.user.pk)

            tipousuario = Permisos.objects.filter(usuid=usuario, TipoPermiso='SUPERADMIN').exists()
            if tipousuario is False:
                verificacion = Poblacion.objects.filter(IdPoblacion=identificador).exists()
                if verificacion is True:
                    tipo = Poblacion.objects.get(IdPoblacion=identificador)
                    tipo.delete()
                    messages.add_message(request, messages.INFO, 'Tipo de poblacion eliminado correctamente')
                    return HttpResponseRedirect(reverse('usuarios:perfil'))
                else:
                    messages.add_message(request, messages.ERROR, 'El tipo de poblacion no EXISTE')
                    return HttpResponseRedirect(reverse('usuarios:perfil'))

            else:
                messages.add_message(request, messages.ERROR,
                                     'Su usuario no tiene los permiso de acceso a esta seccion')
                return HttpResponseRedirect(reverse('usuarios:perfil'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class CambiarContraUsuario(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/cambiarcontrasena.html'

    def get(self, request, usuid):
        try:
            print(usuid)
            usuario2 = Usuario.objects.get(usuid=request.user.pk)
            usuario = Usuario.objects.filter(usuid=usuid)
            tipousuario = Permisos.objects.filter(usuid=usuario2, TipoPermiso='AccessPanel').exists()

            if tipousuario is True:
                return render(request, self.template_name, {'usuario': usuario})

            else:
                messages.add_message(request, messages.ERROR,
                                     'Su usuario no tiene los permiso de acceso a esta seccion')
                return HttpResponseRedirect(reverse('usuarios:inicio'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request, usuid):
        try:
            contrasena = request.POST.get("contrasena","")
            recontrasena = request.POST.get("recontrasena","")

            if contrasena == recontrasena:
                user = User.objects.get(id=usuid)
                user.set_password(contrasena)
                user.save()
                messages.add_message(request, messages.INFO, 'la conseña se cambio correctamente')
                return HttpResponseRedirect(reverse('usuarios:paneladmin'))

            else:
                messages.add_message(request, messages.INFO, 'Las contraseñas no coinciden')
                return HttpResponseRedirect(reverse('usuarios:paneladmin'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

class CambioEstado(LoginRequiredMixin, View):
    login_url = '/'
    form_class = CambioFormEstado
    vervivienda = VisualizarVivienda
    template_name = 'usuarios/cambioestado.html'
    predio = VisualizarVivienda

    def get(self, request, IdVivienda):
        try:
            vivienda = Vivienda.objects.get(IdVivienda=IdVivienda)
            form = self.form_class(instance=vivienda)
            matricula = IdVivienda
            sector = vivienda.Direccion
            casa = vivienda.NumeroCasa
            estado = vivienda.EstadoServicio
            infovivienda = matricula + " " + sector + " " + casa + " " + estado
            usuario = Usuario.objects.get(usuid=request.user.pk)
            tipousuario = Permisos.objects.filter(usuid=usuario, TipoPermiso='NOACFA').exists()
            if tipousuario is False:
                return render(request, self.template_name, {'matricula': matricula,
                                                            'informacion': infovivienda, 'form': form})
            else:
                messages.add_message(request, messages.ERROR,
                                     'Su usuario no tiene los permiso de acceso a esta seccion')
                return HttpResponseRedirect(reverse('usuarios:inicio'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request, IdVivienda):
        try:
            estado = request.POST.get("EstadoServicio")

            if estado == "Operativo":
                vivienda = Vivienda.objects.get(IdVivienda=IdVivienda)
                vivienda.EstadoServicio = "Operativo"
                vivienda.save()
                estadocuenta = EstadoCuenta.objects.get(IdVivienda=IdVivienda)
                estadocuenta.Estado = "Operativo"
                estadocuenta.save()
                ver = self.predio()
                messages.add_message(request, messages.INFO,
                                     'El estado del servicio se modifico correctamente a OPERATIVO')
                ejecutar = ver.get(request, IdVivienda)
                return ejecutar

            elif estado == "Suspendido":
                vivienda = Vivienda.objects.get(IdVivienda=IdVivienda)
                vivienda.EstadoServicio = "Suspendido"
                vivienda.save()
                estadocuenta = EstadoCuenta.objects.get(IdVivienda=IdVivienda)
                estadocuenta.Estado = "Suspendido"
                estadocuenta.save()
                ver = self.predio()
                messages.add_message(request, messages.INFO,
                                     'El estado del servicio se modifico correctamente a SUSPENDIDO')
                ejecutar = ver.get(request, IdVivienda)
                return ejecutar

            elif estado == "Retirado":
                vivienda = Vivienda.objects.get(IdVivienda=IdVivienda)
                vivienda.EstadoServicio = "Retirado"
                vivienda.save()
                estadocuenta = EstadoCuenta.objects.get(IdVivienda=IdVivienda)
                estadocuenta.Estado = "Retirado"
                estadocuenta.save()
                ver = self.predio()
                messages.add_message(request, messages.INFO,
                                     'El estado del servicio se modifico correctamente a RETIRADO')
                ejecutar = ver.get(request, IdVivienda)
                return ejecutar

            else:
                vivienda = Vivienda.objects.get(IdVivienda=IdVivienda)
                vivienda.EstadoServicio = "Mantenimiento"
                vivienda.save()
                estadocuenta = EstadoCuenta.objects.get(IdVivienda=IdVivienda)
                estadocuenta.Estado = "Mantenimiento"
                estadocuenta.save()
                ver = self.predio()
                messages.add_message(request, messages.INFO,
                                     'El estado del servicio se modifico correctamente a MANTENIMIENTO')
                ejecutar = ver.get(request, IdVivienda)
                return ejecutar

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class CierreFinanciero(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/cierrefinanciero.html'

    def get(self, request):
        try:
            usuario = Usuario.objects.get(usuid=request.user.pk)
            idacueducto = usuario.IdAcueducto
            pingregos = Cierres.objects.filter(IdAcueducto=idacueducto).order_by("-IdCierre")

            tipousuario = True
            if tipousuario is True:
                return render(request, self.template_name, {
                    'cierres': pingregos,
                })

            else:
                messages.add_message(request, messages.ERROR,'Su usuario no tiene los permiso de acceso a esta seccion')
                return HttpResponseRedirect(reverse('usuarios:inicio'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request):
        try:
            usuario = Usuario.objects.get(usuid=request.user.pk)
            idacueducto = usuario.IdAcueducto
            acueducto = Acueducto.objects.get(IdAcueducto=idacueducto)
            # entradas de la vista
            username1 = request.POST.get("username", "")
            password1 = request.POST.get("password", "")
            ingresos = request.POST.get("ingresos", "")
            egresos = request.POST.get("egresos", "")
            presupuesto = request.POST.get("presupuesto", "")
            periodo = request.POST.get("periodo", "")
            ano = request.POST.get("ano", "")
            # mensualidades:
            fechaexp = (datetime.today())
            ciclo = fechaexp.month
            ano1 = fechaexp.year
            # Los argumentos serán: Año, Mes, Día, Hora, Minutos, Segundos, Milisegundos.
            new_date = datetime(ano1, ciclo, 1, 1, 00, 00, 00000)
            new_date2 = datetime(ano1, ciclo, 30, 23, 59, 59, 00000)
            pagos2 = Pagos.objects.filter(FechaPago__gte=new_date, FechaPago__lte=new_date2).all()
            gastosaprobados = SolicitudGastos.objects.filter(Fecha__gte=new_date, Fecha__lte=new_date2,
                                                             Estado=ESTADO2).all()
            gasto4 = 0
            for i in gastosaprobados:
                valor = int(i.Valor)
                gasto4 += valor

            pago0 = 0
            for i in pagos2:
                valor = i.ValorPago
                pago0 += int(valor)

            filtro = Cierres.objects.filter(Ciclo=periodo, Ano=ano).exists()
            usuario = auth.authenticate(username=username1, password=password1)
            if usuario is not None and usuario.is_active:
                if filtro is False:
                    if pago0 == int(ingresos) and gasto4 == int(egresos):
                        cierre = Cierres(Ingresos=ingresos, Gastos=egresos, Presupuesto=presupuesto, Ciclo=periodo,
                                         Ano=ano, NoRecaudo=usuario, IdAcueducto=acueducto)
                        cierre.save()
                        messages.add_message(request, messages.INFO, 'El cierre se efectuo correctamente')
                        return HttpResponseRedirect(reverse('usuarios:cierrefinanciero'))

                    else:
                        messages.add_message(request, messages.ERROR,
                                             'Los valores no coinciden con los valores calculados por el sistema')
                        return HttpResponseRedirect(reverse('usuarios:cierrefinanciero'))

                else:
                    messages.add_message(request, messages.ERROR, 'El periodo ingresado ya esta registrado')
                    return HttpResponseRedirect(reverse('usuarios:cierrefinanciero'))

            else:
                messages.add_message(request, messages.ERROR, 'Credenciales incorrectas')
                return HttpResponseRedirect(reverse('usuarios:cierrefinanciero'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class ReporteCierre(LoginRequiredMixin, View):
    login_url = '/'

    def get(self, request):
        try:
            cierres = Cierres.objects.all()
            sfecha = (datetime.today())
            wb = Workbook()
            ws = wb.active
            ws.title = 'Reporte de cierres'
            ws['A1'] = 'Id Cierre'
            ws['B1'] = 'Ingresos'
            ws['C1'] = 'Egresos'
            ws['D1'] = 'Saldo'
            ws['E1'] = 'Periodo'
            ws['F1'] = 'Año'
            ws['G1'] = 'Fecha de cierre'
            ws['H1'] = 'Usuario'
            cont = 2

            for cierre in cierres:
                ws.cell(row=cont, column=1).value = cierre.IdCierre
                ws.cell(row=cont, column=2).value = cierre.Ingresos
                ws.cell(row=cont, column=3).value = cierre.Gastos
                ws.cell(row=cont, column=4).value = cierre.Presupuesto
                ws.cell(row=cont, column=5).value = cierre.Ciclo
                ws.cell(row=cont, column=6).value = cierre.Ano
                ws.cell(row=cont, column=7).value = cierre.Fecha
                ws.cell(row=cont, column=8).value = cierre.NoRecaudo

                cont += 1

            archivo_predios = "ReporteCierresPeriodicos" + str(sfecha) + ".xlsx"
            response = HttpResponse(content_type="application/ms-excel")
            content = "attachment; filename = {0}".format(archivo_predios)
            response['Content-Disposition'] = content
            wb.save(response)
            return response

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class ReporteGastos(LoginRequiredMixin, View):
    login_url = '/'

    def get(self, request):
        try:
            cierres = SolicitudGastos.objects.all()
            sfecha = (datetime.today())
            wb = Workbook()
            ws = wb.active
            ws.title = 'Reporte de gastos'
            ws['A1'] = 'Id gasto'
            ws['B1'] = 'Usuario registro'
            ws['C1'] = 'Descripcion'
            ws['D1'] = 'Tipo de solicitud'
            ws['E1'] = 'Valor'
            ws['F1'] = 'Estado'
            ws['G1'] = 'Fecha'
            ws['H1'] = 'AreaResponsable'
            cont = 2

            for cierre in cierres:
                ws.cell(row=cont, column=1).value = cierre.IdSoGa
                ws.cell(row=cont, column=2).value = str(cierre.IdUsuario)
                ws.cell(row=cont, column=3).value = cierre.Descripcion
                ws.cell(row=cont, column=4).value = cierre.TipoSolicitud
                ws.cell(row=cont, column=5).value = cierre.Valor
                ws.cell(row=cont, column=6).value = cierre.Estado
                ws.cell(row=cont, column=7).value = cierre.Fecha
                ws.cell(row=cont, column=8).value = cierre.AreaResponsable

                cont += 1

            archivo_predios = "Reportegastos" + str(sfecha) + ".xlsx"
            response = HttpResponse(content_type="application/ms-excel")
            content = "attachment; filename = {0}".format(archivo_predios)
            response['Content-Disposition'] = content
            wb.save(response)
            return response

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class ImprimirSoporteP(LoginRequiredMixin, View):
    login_url = '/'

    def get(self, request, IdPago):
        try:
            # pago
            usuario = Usuario.objects.get(usuid=request.user.pk)
            telefono = str(usuario.IdAcueducto.Telefono)
            pago = Pagos.objects.get(IdPago=IdPago)
            idfactura = pago.IdFactura
            idpago = str(pago.IdPago)
            fecha1 = pago.FechaPago
            fecha = str(fecha1.year) + '-' + str(fecha1.month) + '-' + str(fecha1.day) + ' ' + str(
                fecha1.hour) + ':' + str(fecha1.minute)
            valorpago = str(pago.ValorPago)
            resta = str(pago.resta)
            # factura
            factura = Facturas.objects.get(IdFactura=idfactura.pk)
            idfac = str(factura.IdFactura)
            periodo = factura.IdCiclo.Nombre
            referencia = str(factura.IdVivienda)
            totalfactura = str(factura.Total)
            # Impresora
            nombreImpresora = "termica3"
            conector = ConectorV3()
            conector.Iniciar()
            conector.DeshabilitarElModoDeCaracteresChinos()
            conector.EstablecerAlineacion(ALINEACION_CENTRO)
            conector.TextoSegunPaginaDeCodigos(2, "cp850", "¡AUECAAC ESP!\n")
            conector.EscribirTexto("NIT 900.017.239-2\n")
            conector.EstablecerEnfatizado(True)
            conector.EscribirTexto("COMPROBANTE DE PAGO\n")
            conector.EstablecerEnfatizado(False)
            conector.Feed(1)
            conector.EstablecerAlineacion(ALINEACION_IZQUIERDA)
            conector.EscribirTexto("Numero de transaccion: ")
            conector.EscribirTexto(idpago + "\n")
            conector.EscribirTexto("Numero de factura: ")
            conector.EscribirTexto(idfac + "\n")
            conector.EscribirTexto("Fecha de pago: ")
            conector.EscribirTexto(fecha + "\n")
            conector.EscribirTexto("Punto de pago: Oficina principal\n")
            conector.EscribirTexto("Periodo de pago: " + periodo + "\n")
            conector.EscribirTexto("Matricula: \n")
            conector.EscribirTexto(referencia + "\n")
            conector.Feed(1)
            conector.EstablecerAlineacion(ALINEACION_CENTRO)
            conector.EscribirTexto("_ _ _ _ _ _ _ _ _ _ _ _ _ _ _ _\n")
            conector.EstablecerEnfatizado(True)
            conector.EscribirTexto("Valor pagado: ")
            conector.EstablecerEnfatizado(False)
            conector.EscribirTexto("$ " + valorpago + "\n")
            conector.EscribirTexto("_ _ _ _ _ _ _ _ _ _ _ _ _ _ _ _\n")
            conector.EstablecerEnfatizado(True)
            conector.EscribirTexto("Para cualquier acto de reclamacion, debera presentar este soporte de pago\n")
            conector.EstablecerEnfatizado(False)
            conector.Feed(1)
            conector.EscribirTexto("_ _ _ _ _ _ _ _ _ _ _ _ _ _ _ _\n")
            conector.Feed(1)
            conector.EstablecerAlineacion(ALINEACION_CENTRO)
            conector.Corte(1)
            conector.Pulso(48, 60, 120)
            respuesta = conector.imprimirEn(nombreImpresora)
            if respuesta is True:
                messages.add_message(request, messages.INFO, 'Comprobante de pago, se genero correctamente')
                return HttpResponseRedirect(reverse('usuarios:inicio'))
            else:
                messages.add_message(request, messages.ERROR, respuesta)
                return HttpResponseRedirect(reverse('usuarios:inicio'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class PazSalvo(LoginRequiredMixin, View):
    login_url = '/'
    vervivienda = VisualizarVivienda

    def get(self, request, matricula):
        try:
            idvivienda = matricula
            vivienda = Vivienda.objects.get(IdVivienda=idvivienda)
            estado = vivienda.EstadoServicio
            idpropietario = vivienda.IdPropietario
            titular = Propietario.objects.get(IdPropietario=idpropietario.pk)
            nombrecompleto = str(titular.Nombres) + ' ' + str(titular.Apellidos)
            sector = vivienda.Direccion
            casa = vivienda.NumeroCasa
            direccion = sector + ' Casa No ' + casa
            now = datetime.now()
            dia = now.day
            mes = now.month
            anio = now.year

            if 1 == 1:
                estadocuenta = EstadoCuenta.objects.get(IdVivienda=idvivienda)
                valor = estadocuenta.Valor
                if valor <= 0:
                    wb = openpyxl.load_workbook('static/Formatos/E8CPS.xlsx')
                    ws = wb.active
                    # primer mensaje
                    ws['U9'] = idvivienda
                    ws['A10'] = nombrecompleto
                    ws['A11'] = direccion
                    ws[
                        'A16'] = 'no aplica'
                    ws['A22'] = 'El presente certificado se expide por solicitud del interesado a los ' + str(
                        dia) + ' días del mes ' + str(mes) + ' del año ' + str(anio)
                    if estado == 'Operativo':
                        ws[
                            'A26'] = 'Como el predio tiene la matricula activa, el presente certificado ' \
                                     'solo es valido por 30 dias a partir de la fecha de expedicion.'
                    else:
                        ws['A26'] = ' '
                    ws.title = idvivienda
                    archivo_predios = "paz y salvo " + str(idvivienda) + ".xlsx"
                    response = HttpResponse(content_type="application/ms-excel")
                    content = "attachment; filename = {0}".format(archivo_predios)
                    response['Content-Disposition'] = content
                    wb.save(response)
                    return response
                else:
                    messages.add_message(request, messages.ERROR,
                                         'El predio no se encuentra a paz y salvo por conceptos de acueducto')
                    ver = self.vervivienda()
                    ejercutar = ver.get(request, idvivienda)
                    return ejercutar

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

class CobroRecargo(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/cobrorecargo.html'

    def get(self, request):
        try:
            estadoscuenta = EstadoCuenta.objects.all()
            cont = 0
            for i in estadoscuenta:
                valor = int(i.Valor)
                if valor >= 17000:
                    cont += 1

            return render(request, self.template_name, {
                'cont': cont})

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request):
        try:
            usuario = Usuario.objects.get(usuid=request.user.pk)
            idacueducto = usuario.IdAcueducto
            acueducto = Acueducto.objects.get(IdAcueducto=idacueducto)
            idtarifa = acueducto.IdTarifa
            tarifa = Tarifa.objects.get(IdTarifa=idtarifa)
            recargo = tarifa.Recargo

            estadoscuenta = EstadoCuenta.objects.all()
            cont = 0
            for i in estadoscuenta:
                valor = i.Valor
                if valor >= 17000:
                    cont += 1
                    estadoscu = EstadoCuenta.objects.get(IdEstadoCuenta=i.pk)
                    estadoscu.Valor += int(recargo)
                    estadoscu.save()

            messages.add_message(request, messages.INFO, 'Se genero el cobro de los recargos correctamente')
            return HttpResponseRedirect(reverse('usuarios:estadoscuenta'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class VerFactura(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/verfactura.html'
    imptiquet = ImprimirSoporteP

    def get(self, request):
        try:
            numerofactura = request.GET.get('factura', ' ')
            usuario = Usuario.objects.get(usuid=request.user.pk)
            idacueducto = usuario.IdAcueducto
            anulada = Facturas.objects.filter(IdFactura=numerofactura, Estado=FA,IdAcueducto=idacueducto).exists()
            paga = Facturas.objects.filter(IdFactura=numerofactura, Estado=FP,IdAcueducto=idacueducto).exists()
            vencida = Facturas.objects.filter(IdFactura=numerofactura, Estado=FV,IdAcueducto=idacueducto).exists()
            validacion = Facturas.objects.filter(IdFactura=numerofactura,IdAcueducto=idacueducto).exists()
            if validacion is True:
                factura = Facturas.objects.get(IdFactura=numerofactura)
                idvivienda = factura.IdVivienda
                concepfac = FacturasConceptos.objects.filter(IdFactura=numerofactura)
                conceptosfac = ConceptosFacturados.objects.filter(IdVivienda=idvivienda)
                vivienda = Vivienda.objects.get(IdVivienda=idvivienda.pk)
                matricula = vivienda.IdVivienda
                sector = vivienda.Direccion
                casa = vivienda.NumeroCasa
                piso = vivienda.Piso
                fe = factura.FechaExpe
                fl = factura.FechaLimite
                total = factura.Total
                ciclo = factura.IdCiclo
                estadofactura = factura.Estado
                return render(request, self.template_name, {
                        'estadoscuenta': total,
                        'conceptos': concepfac,
                        'conceptosfac':conceptosfac,
                        'factura': numerofactura,
                        'fe': fe,
                        'fl': fl,
                        'total': total,
                        'estadofactura': estadofactura,
                        'matricula': matricula,
                        'casa': casa,
                        'sector': sector,
                        'piso': piso,
                        'ciclo': ciclo,
                        'anulada': anulada,
                        'paga': paga,
                        'vencida': vencida,
                    })

            else:
                messages.add_message(request, messages.ERROR,'El numero de factura ingresado, no existe o no le corresponde a su empresa')
                return HttpResponseRedirect(reverse('usuarios:inicio'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request):
        try:
            usuario = Usuario.objects.get(usuid=request.user.pk)
            idacueducto = usuario.IdAcueducto
            acueducto = Acueducto.objects.get(IdAcueducto=idacueducto)
            numerofactura = request.POST.get("factura", "")
            tipo = request.POST.get("tipo")
            abono = request.POST.get("concepto")
            factura = Facturas.objects.get(IdFactura=numerofactura)
            idvivienda = factura.IdVivienda
            mensualidad = factura.periodofacturado
            formato = "%Y"
            s = (datetime.today())
            ano = s.strftime(formato)
            s = (datetime.today())
            fecha = s + timedelta(days=2)

            if tipo == 'Total':
                relacion = FacturasConceptos.objects.get(IdFactura=numerofactura)
                idconcepto = relacion.IdConcepto
                concepto = ConceptosFacturados.objects.get(IdRegistro=idconcepto.pk)
                suma = concepto.Total
                concepto.Estado = 'Pago'
                concepto.save()

                pago = Pagos(IdFactura=factura, Ano=ano, ValorPago=suma, Descripcion=mensualidad,
                             Efectivo=0, Devuelta=0, IdUsuario=usuario, IdVivienda=idvivienda, resta=0, IdAcueducto=acueducto)
                pago.save()

                idpago = pago.IdPago

                suspencion = OrdenesTrabajo.objects.filter(IdVivienda=idvivienda, Estado='Pendiente').exists()
                if suspencion is True:
                    suspencion2 = OrdenesTrabajo.objects.get(IdVivienda=idvivienda, Estado='Pendiente')
                    suspencion2.Estado = 'Anulada'
                    suspencion2.save()

                cambiofactura = Facturas.objects.get(IdFactura=numerofactura)
                cambiofactura.Estado = FP
                cambiofactura.save()

                descripcion = 'Se registra pago: ' + str(idpago) + ' Factura: ' + str(
                    numerofactura) + ' Matricula: ' + str(idvivienda) + ' Valor: $' + str(suma)

                novedad = Novedades(Descripcion=descripcion, TipoNovedad='Pago', usuario=usuario, matricula=idvivienda,IdAcueducto=acueducto)
                novedad.save()
                tiquet2 = self.imptiquet()
                ejecutar = tiquet2.get(request, idpago)
                return ejecutar

            elif tipo == 'Abono':
                relacion = FacturasConceptos.objects.get(IdFactura=numerofactura)
                idconcepto = relacion.IdConcepto
                concepto = ConceptosFacturados.objects.get(IdRegistro=idconcepto.pk)
                total = concepto.Total
                concepto.Estado = 'Abono'
                concepto.save()
                periodo = concepto.Periodo

                resta = int(total) - int(abono)

                pago = Pagos(IdFactura=factura, Ano=ano, ValorPago=abono, Descripcion=mensualidad,
                             Efectivo=0, Devuelta=0,
                             IdUsuario=usuario, IdVivienda=idvivienda, resta=resta, IdAcueducto=acueducto)
                pago.save()
                idpago = pago.IdPago

                suspencion = OrdenesTrabajo.objects.filter(IdVivienda=idvivienda, Estado='Pendiente').exists()
                if suspencion is True:
                    suspencion2 = OrdenesTrabajo.objects.get(IdVivienda=idvivienda, Estado='Pendiente')
                    suspencion2.Estado = 'Anulada'
                    suspencion2.save()

                cambiofactura = Facturas.objects.get(IdFactura=numerofactura)
                cambiofactura.Estado = FP
                cambiofactura.save()

                concepto1 = Conceptos(Tipo='Saldo P', Observacion=str(idconcepto) + ' - ' + str(periodo) + ' - Saldo pendiente: - $' + str(resta),
                                      Estado='Sin facturar', Valor=resta, IdVivienda=idvivienda)
                concepto1.save()

                descripcion = 'Se registra pago: ' + str(idpago) + ' Factura: ' + str(
                    numerofactura) + ' Matricula: ' + str(idvivienda) + ' Valor: $' + str(abono)
                novedad = Novedades(Descripcion=descripcion, TipoNovedad='Pago', usuario=usuario, matricula=idvivienda, IdAcueducto=acueducto)
                novedad.save()

                tiquet2 = self.imptiquet()
                ejecutar = tiquet2.get(request, idpago)
                return ejecutar

            else:
                messages.add_message(request, messages.INFO, 'No se puede realizar la transaccion')
                return HttpResponseRedirect(reverse('usuarios:inicio'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class AnularPago(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/anularpago.html'

    def get(self, request):
        try:
            usuario = 0
            if usuario == 0:
                return render(request, self.template_name)

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request):
        try:
            usuario = Usuario.objects.get(usuid=request.user.pk)
            comprobante = request.POST.get("Numero")
            pago = Pagos.objects.get(IdPago=comprobante)
            matricula = pago.IdVivienda
            valorpagado = pago.ValorPago
            estadocuenta = EstadoCuenta.objects.get(IdVivienda=matricula)
            factura = Facturas.objects.get(IdFactura=pago.IdFactura.pk)
            if int(comprobante) >= 1:
                factura.Estado = 'Emitida'
                factura.save()
                valor = estadocuenta.Valor + int(valorpagado)
                estadocuenta.Valor = valor
                estadocuenta.save()
                descripcion = 'Se anula pago no: ' + str(comprobante) + ' Matricula: ' + str(
                    pago.IdVivienda) + ' valor: ' + str(pago.ValorPago)
                novedad = Novedades(Descripcion=descripcion, matricula=matricula, usuario=usuario, TipoNovedad='Pago anulado')
                novedad.save()
                pago1 = Pagos.objects.get(IdPago=comprobante)
                pago1.delete()
                messages.add_message(request, messages.INFO, 'El pago se anulo correctamente')
                return HttpResponseRedirect(reverse('usuarios:inicio'))

            else:
                messages.add_message(request, messages.ERROR, 'No se pudo anular el pago')
                return HttpResponseRedirect(reverse('usuarios:inicio'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class Matriculas(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/matriculas.html'

    def get(self, request):
        try:
            # lista de predios
            pnv = Vivienda.objects.filter(Direccion='Pasonivel Viejo').count()
            pnd = Vivienda.objects.filter(Direccion='Pasonivel Destapada').count()
            cc = Vivienda.objects.filter(Direccion='Caimalito Centro').count()
            bn = Vivienda.objects.filter(Direccion='Barrio Nuevo').count()
            vj = Vivienda.objects.filter(Direccion='20 de julio').count()
            hd = Vivienda.objects.filter(Direccion='Hacienda').count()
            carb = Vivienda.objects.filter(Direccion='Carbonera').count()

            # disponible
            epnv = AsignacionBloque.objects.filter(Bloque='PNV', Estado='Sin asignar').count()
            epnd = AsignacionBloque.objects.filter(Bloque='PND', Estado='Sin asignar').count()
            ecc = AsignacionBloque.objects.filter(Bloque='CC', Estado='Sin asignar').count()
            ebn = AsignacionBloque.objects.filter(Bloque='BN', Estado='Sin asignar').count()
            ehd = AsignacionBloque.objects.filter(Bloque='HD', Estado='Sin asignar').count()
            evj = AsignacionBloque.objects.filter(Bloque='VJ', Estado='Sin asignar').count()
            car = AsignacionBloque.objects.filter(Bloque='CA', Estado='Sin asignar').count()

            # filtros
            fpnv = AsignacionBloque.objects.filter(Bloque='PNV', Estado='Sin asignar')
            fpnd = AsignacionBloque.objects.filter(Bloque='PND', Estado='Sin asignar')
            fcc = AsignacionBloque.objects.filter(Bloque='CC', Estado='Sin asignar')
            fbn = AsignacionBloque.objects.filter(Bloque='BN', Estado='Sin asignar')
            fhd = AsignacionBloque.objects.filter(Bloque='HD', Estado='Sin asignar')
            fvj = AsignacionBloque.objects.filter(Bloque='VJ', Estado='Sin asignar')
            fcar = AsignacionBloque.objects.filter(Bloque='CA', Estado='Sin asignar')

            usuario = 0
            if usuario == 0:
                return render(request, self.template_name,
                              {'pnv': pnv, 'pnd': pnd, 'cc': cc, 'bn': bn, 'vj': vj, 'hd': hd,
                               'epnv': epnv, 'epnd': epnd, 'ecc': ecc, 'ebn': ebn, 'ehd': ehd, 'evj': evj,
                               'fhd': fhd, 'fpnv': fpnv, 'fpnd': fpnd, 'fvj': fvj, 'fbn': fbn, 'fcc': fcc,
                               'car':car,'fcar':fcar, 'carb':carb
                               })
        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class Bloque(LoginRequiredMixin, View):
    login_url = '/'

    def get(self):
        sector = 'Pasonivel Viejo'
        predios = Vivienda.objects.filter(Direccion=sector)
        suma = 0
        for i in predios:
            idvivienda = i.IdVivienda
            estadocuenta = EstadoCuenta.objects.get(IdVivienda=idvivienda)
            idestadocuenta = estadocuenta.IdEstadoCuenta
            buscar = AsignacionBloque.objects.get(Matricula=idvivienda)
            buscar.Estado = 'Asignada'
            buscar.Estadocuenta = idestadocuenta
            buscar.save()
            suma += 1

        return suma


class AsignarCargo(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/asignacioncargo.html'
    form_class = CobroMatriculaForm
    vizualizarv = VisualizarVivienda

    def get(self, request, matricula):
        try:
            form = self.form_class
            vivienda = Vivienda.objects.get(IdVivienda=matricula)
            direccion = vivienda.Direccion + ' Cs ' + vivienda.NumeroCasa + ' Piso ' + vivienda.Piso
            bloque = AsignacionBloque.objects.get(Matricula=matricula)
            fecha = bloque.Fecha
            estadocuenta = bloque.Estadocuenta
            usuario1 = 0
            if usuario1 == 0:
                return render(request, self.template_name,
                              {'form': form, 'matricula': matricula, 'direccion': direccion, 'fecha': fecha,
                               'cuenta': estadocuenta})

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request, matricula):
        try:
            usuario = Usuario.objects.get(usuid=request.user.pk)
            idacueducto = usuario.IdAcueducto
            idvivienda = matricula
            descripcion = 'Cargo por conexion'
            estado = 'Pendiente'
            idvalor = request.POST.get("IdValor")
            cantcuotas = request.POST.get("CantCuotas")
            cuotaspendientes = cantcuotas
            vivienda = Vivienda.objects.filter(IdVivienda=idvivienda).exists()
            vivienda2 = Vivienda.objects.get(IdVivienda=idvivienda)
            valor = ValorMatricula.objects.get(IdValor=idvalor)
            valorpendiente = valor.Valor
            cuota = int(valorpendiente)
            if vivienda is True:
                matricula = CobroMatricula(Descripcion=descripcion, IdVivienda=vivienda2, Estado=estado, IdValor=valor,
                                           CantCuotas=cantcuotas,
                                           CuotasPendientes=cuotaspendientes, ValorPendiente=valorpendiente,
                                           Cuota=cuota,IdAcueducto=idacueducto)
                matricula.save()
                ver = self.vizualizarv()
                messages.add_message(request, messages.INFO, 'La informacion del predio se agrego correctamente')
                ejercutar = ver.get(request, idvivienda)
                return ejercutar

            else:
                messages.add_message(request, messages.ERROR, 'ERROR')
                return HttpResponseRedirect(reverse('usuarios:matriculas'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class Creditos(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/creditos.html'

    def get(self, request):
        try:
            usuario = Usuario.objects.get(usuid=request.user.pk)
            idacueducto = usuario.IdAcueducto
            creditos = Credito.objects.filter(IdAcueducto=idacueducto)
            vigentes = Credito.objects.filter(Estado='Vigente', IdAcueducto=idacueducto).count()
            creditosv = Credito.objects.filter(Estado='Vigente', IdAcueducto=idacueducto)
            creditosp = Credito.objects.filter(Estado='Pagado', IdAcueducto=idacueducto)
            pagados = Credito.objects.filter(Estado='Pagado', IdAcueducto=idacueducto).count()
            deudatotal = 0
            for i in creditos:
                valor = i.ValorInicial
                deudatotal += int(valor)

            deudapendiente = 0
            for i in creditos:
                valor = i.ValorPendiente
                deudapendiente += int(valor)

            usuario1 = 0
            if usuario1 == 0:
                return render(request, self.template_name, {'vigentes': vigentes,
                                                            'creditosv': creditosv, 'creditosp': creditosp,
                                                            'deudatotal': deudatotal,
                                                            'deudapendiente': deudapendiente, 'pagados': pagados})

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class RegistroCredito(LoginRequiredMixin, View):
    login_url = '/'
    form_class = FormRegistroCredito
    template_name = 'usuarios/registrocredito.html'

    def get(self, request):
        try:
            form = self.form_class()
            return render(request, self.template_name,
                          {
                              'form': form
                          })

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request):
        try:
            form = self.form_class(request.POST)

            if form.is_valid():
                form.save()
                messages.add_message(request, messages.INFO, 'El credito se agrego correctamente')
                return HttpResponseRedirect(reverse('usuarios:credito'))

            else:
                messages.add_message(request, messages.ERROR, 'No se pudo agregar la informacion al sistema')
                return HttpResponseRedirect(reverse('usuarios:credito'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class RegistroProveedor(LoginRequiredMixin, View):
    login_url = '/'
    form_class = FormRegistroProveedor
    template_name = 'usuarios/registroproveedor.html'

    def get(self, request, *args, **kwargs):
        try:
            form = self.form_class()
            return render(request, self.template_name,
                          {
                              'form': form
                          })

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request, *args):
        try:
            form = self.form_class(request.POST)

            if form.is_valid():
                form.save()
                messages.add_message(request, messages.INFO, 'La informacion del proveedor se agrego correctamente')
                return HttpResponseRedirect(reverse('usuarios:controlpresupuestal'))

            else:
                messages.add_message(request, messages.ERROR, 'No se pudo agregar la informacion al sistema')
                return HttpResponseRedirect(reverse('usuarios:controlpresupuestal'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class VerCredito(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/vercredito.html'

    def get(self, request, idcredito):
        try:
            credito = Credito.objects.filter(IdCredito=idcredito)
            credito2 = Credito.objects.get(IdCredito=idcredito)
            estado = credito2.Estado
            verificarpagos = SolicitudGastos.objects.filter(NumeroFactura=idcredito)
            return render(request, self.template_name, {'credito': credito, 'pagos': verificarpagos, 'estado': estado})

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request, idcredito):
        try:
            usuario = Usuario.objects.get(usuid=request.user.pk)
            idacueducto = usuario.IdAcueducto
            acueducto = Acueducto.objects.get(IdAcueducto=idacueducto)
            valor = request.POST.get("valor")
            credito = Credito.objects.get(IdCredito=idcredito)
            numcuota = credito.CuotasPendiente
            nombrecredito = credito.NombreCredito
            mensaje = 'Cuota no' + str(numcuota) + '|' + str(nombrecredito)
            savegasto = SolicitudGastos(Descripcion=mensaje, TipoSolicitud='Pago de creditos', Valor=valor,
                                        Estado='Pendiente',
                                        AreaResponsable='Area Administrativa', NumeroFactura=idcredito,
                                        IdUsuario=usuario,IdAcueducto=acueducto)
            savegasto.save()
            cuotamenos = int(numcuota) - 1
            valorpen = int(credito.ValorPendiente) - int(valor)
            if cuotamenos <= 0:
                credito.CuotasPendiente = str(cuotamenos)
                credito.ValorPendiente = str(valorpen)
                credito.Estado = 'Pagado'
                credito.save()
                messages.add_message(request, messages.INFO, 'El credito fue pagado con exito')
                return HttpResponseRedirect(reverse('usuarios:credito'))

            else:
                credito.CuotasPendiente = str(cuotamenos)
                credito.ValorPendiente = str(valorpen)
                credito.save()
                messages.add_message(request, messages.INFO, 'La cuota se registro correctamente')
                return HttpResponseRedirect(reverse('usuarios:credito'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class ReporteRetiro(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/retiro.html'

    def get(self, request, matricula):
        try:
            usuario = 0
            vivienda = Vivienda.objects.get(IdVivienda=matricula)
            matricula1 = vivienda.IdVivienda
            if usuario == 0:
                return render(request, self.template_name, {'idvivienda': matricula1})

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request, matricula):
        try:
            usuarios = Usuario.objects.get(usuid=request.user.pk)
            idacueducto = usuarios.IdAcueducto
            acueducto = Acueducto.objects.get(IdAcueducto=idacueducto)
            Descripcion = request.POST.get("descripcion")
            usuario =0
            if usuario == 0:
                vivienda = Vivienda.objects.get(IdVivienda=matricula)
                registrarnovedad = Novedades(matricula=vivienda,Descripcion=Descripcion, TipoNovedad='Retiro',usuario=usuarios,IdAcueducto=acueducto)
                vivienda.EstadoServicio = 'Retirado'
                vivienda.save()
                registrarnovedad.save()
                messages.add_message(request, messages.INFO, 'La novedad se registro correctamente')
                return HttpResponseRedirect(reverse('usuarios:inicio'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class Consumo(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/Consumos.html'

    def get(self, request):
        try:
            usuario2 = 0
            usuario = Usuario.objects.get(usuid=request.user.pk)
            idacueducto = usuario.IdAcueducto
            sinasignar = Medidores.objects.filter(Estado='Sin asignar', IdAcueducto=idacueducto)
            asignado = Asignacion.objects.filter(Estado='Operativo', IdAcueducto=idacueducto)
            contar = Asignacion.objects.filter(Estado='Operativo', IdAcueducto=idacueducto).count()
            vapo = Consumos.objects.filter(IdAcueducto=idacueducto).aggregate(Consumo=Sum('Consumo'))
            suma = vapo['Consumo']
            fechaexp = (datetime.today())
            anio = fechaexp.year
            vapo2 = ConsumosMensual.objects.filter(IdAcueducto=idacueducto, ano=anio).aggregate(Consumo=Sum('Consumo'))
            suma2 = vapo2['Consumo']
            novedades = ConsumosMensual.objects.filter(IdAcueducto=idacueducto).order_by("-IdRegistro")[:1]
            if usuario2 == 0:
                return render(request, self.template_name,{
                    'sinasignar':sinasignar, 'asignado':asignado, 'contar': contar,'suma':suma,'suma2': suma2,'ultimo':novedades,
                })

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class RegistrarConsumo(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/Registrarconsumos.html'

    def get(self, request):
        try:
            usuario = Usuario.objects.get(usuid=request.user.pk)
            idacueducto = usuario.IdAcueducto
            matricula = request.GET.get('matricula')
            vivienda = Vivienda.objects.get(IdVivienda=matricula)
            asignado = Asignacion.objects.filter(IdVivienda=matricula,Estado='Operativo', IdAcueducto=idacueducto).exists()

            if asignado == True:
                return render(request, self.template_name,{'matricula': matricula})
            else:
                messages.add_message(request, messages.ERROR, 'El predio no tiene micromedidor asignado')
                return HttpResponseRedirect(reverse('usuarios:consumos'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

    def post(self, request):
        try:
            usuario = Usuario.objects.get(usuid=request.user.pk)
            idacueducto = usuario.IdAcueducto
            acueducto = Acueducto.objects.get(IdAcueducto=idacueducto)
            lecturaactual = request.POST.get("lectura")
            observacion = request.POST.get("observacion")
            matricula = request.POST.get("matricula")
            medidor = Asignacion.objects.get(IdVivienda=matricula,IdAcueducto=idacueducto,Estado='Operativo')
            idmedidor = medidor.IdMedidor
            vivienda = Vivienda.objects.get(IdVivienda=matricula,IdAcueducto=idacueducto)
            consultarconsumo = Consumos.objects.filter(IdVivienda=matricula,IdAcueducto=idacueducto).exists()

            fechaexp = (datetime.today())
            mes1 = fechaexp.month
            ciclos = Ciclo.objects.get(IdCiclo=mes1)
            mes = ciclos.Nombre
            ano = fechaexp.year

            if consultarconsumo == False:
                consumo1 = Consumos(IdMedidor=idmedidor, IdVivienda=vivienda, Lecturaactual=lecturaactual,Lecturaanterior=0,Consumo=lecturaactual,promedio=lecturaactual,
                                    observaciones=observacion,diasconsumo=30, ano=ano, mes=mes, IdAcueducto=idacueducto)
                consumo1.save()
                messages.add_message(request, messages.INFO, 'La medicion se agrego correctamente')
                return HttpResponseRedirect(reverse('usuarios:consumos'))

            else:
                vapo = Consumos.objects.filter(IdVivienda=matricula,IdAcueducto=idacueducto).aggregate(Consumo=Sum('Consumo'))
                suma = vapo['Consumo']
                cantidadregistros = Consumos.objects.filter(IdVivienda=matricula,IdAcueducto=idacueducto).count()
                consultarconsumo = Consumos.objects.filter(IdVivienda=matricula,IdAcueducto=idacueducto).order_by("-IdRegistro")[:1]
                primerobjeto = consultarconsumo[0]

                lecturaanterior = primerobjeto.Lecturaactual
                consumo = int(lecturaactual) - int(lecturaanterior)
                suma2 = consumo + suma
                promedio = suma2 / (cantidadregistros + 1)
                consumo1 = Consumos(IdMedidor=idmedidor, IdVivienda=vivienda, Lecturaactual=lecturaactual,
                                    Lecturaanterior=lecturaanterior, Consumo=consumo, promedio=int(promedio),
                                    observaciones=observacion, diasconsumo=30, ano=ano, mes=mes,IdAcueducto=idacueducto)
                consumo1.save()

                messages.add_message(request, messages.INFO, 'La medicion se agrego correctamente')
                return HttpResponseRedirect(reverse('usuarios:consumos'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

class VerConsumo(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/consumosuscriptor.html'

    def get(self, request, matricula):
        try:

            usuario = Usuario.objects.get(usuid=request.user.pk)
            tipousuario = True
            vivienda = Vivienda.objects.get(IdVivienda=matricula)
            asignado = Asignacion.objects.get(IdVivienda=matricula,Estado='Operativo')
            idmedidor = asignado.IdMedidor
            medidor = Medidores.objects.get(IdMedidor=idmedidor)
            consumos = Consumos.objects.filter(IdVivienda=matricula).order_by("-IdRegistro")
            vapo = Consumos.objects.filter(IdVivienda=matricula).aggregate(Consumo=Sum('Consumo'))
            totalconsumo = vapo['Consumo']

            if tipousuario is True:
                return render(request, self.template_name, {'medidor': medidor.IdMedidor, 'estado':medidor.Estado,'estado2': asignado.Estado,
                                                            'matricula': matricula, 'consumos': consumos, 'totalconsumo':totalconsumo})
            else:
                messages.add_message(request, messages.ERROR, 'Su usuario no tiene los permisos de acceso a esta '
                                                              'seccion')
                return HttpResponseRedirect(reverse('usuarios:inicio'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

class FacturadorConceptos(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/consumosuscriptor.html'

    def get(self, request):
        try:
            usuario = Usuario.objects.get(usuid=request.user.pk)
            idacueducto = usuario.IdAcueducto
            acueducto = Acueducto.objects.get(IdAcueducto=idacueducto)

            fechaexp = (datetime.today())
            fechalimite = fechaexp + timedelta(days=DIASPARASUSPENCION)
            mes1 = fechaexp.month
            ciclos = Ciclo.objects.get(IdCiclo=mes1)
            mes = ciclos.Nombre
            tipousuario = True
            viviendas = Vivienda.objects.filter(EstadoServicio='Operativo')|Vivienda.objects.filter(EstadoServicio='Suspendido')
            if tipousuario is True:
                for i in viviendas:
                    conceptos = Conceptos.objects.filter(Estado='Sin facturar', IdVivienda=i.IdVivienda).exists()
                    vivienda = Vivienda.objects.get(IdVivienda=i.IdVivienda)
                    if conceptos is True:
                        concepto = Conceptos.objects.filter(Estado='Sin facturar', IdVivienda=i.IdVivienda)

                        aportefijo = 0
                        basico = 0
                        suspencion = 0
                        reconexion = 0
                        subsidio = 0
                        complemen = 0
                        recargo = 0
                        acuerdopago = 0
                        saldopen = 0

                        for j in concepto:
                            if j.Tipo == 'Aporte fijo':
                                aportefijo = j.Valor
                                j.Estado = 'Facturado'
                                j.save()

                            if j.Tipo == 'Basico':
                                basico = j.Valor
                                j.Estado = 'Facturado'
                                j.save()

                            if j.Tipo == 'Suspencion':
                                suspencion = j.Valor
                                j.Estado = 'Facturado'
                                j.save()

                            if j.Tipo == 'Reconexion':
                                reconexion = j.Valor
                                j.Estado = 'Facturado'
                                j.save()

                            if j.Tipo == 'Subsidio':
                                subsidio = j.Valor
                                j.Estado = 'Facturado'
                                j.save()

                            if j.Tipo == 'Consumo complementario':
                                complemen = j.Valor
                                j.Estado = 'Facturado'
                                j.save()

                            if j.Tipo == 'Recargo':
                                recargo = j.Valor
                                j.Estado = 'Facturado'
                                j.save()

                            if j.Tipo == 'Financiacion':
                                acuerdopago = j.Valor
                                j.Estado = 'Facturado'
                                j.save()

                            if j.Tipo == 'Saldo P':
                                saldopen = j.Valor
                                j.Estado = 'Facturado'
                                j.save()

                        valortotal = acuerdopago + saldopen + aportefijo + complemen  + suspencion + reconexion + recargo + basico - subsidio
                        facturarconcepto = ConceptosFacturados(AporteFijo=aportefijo, Suspencion=suspencion,Reconexion=reconexion,CuotaMatricula=0,
                                          Subsidio=subsidio,Estado='Pendiente', Periodo=mes, Complementario= complemen,Basico=basico,
                                          Recargo=recargo,AcuerdoPago=acuerdopago,SaldoAnterior=saldopen,Total=valortotal, IdVivienda=vivienda,IdAcueducto=acueducto)
                        facturarconcepto.save()

                messages.add_message(request, messages.INFO, 'La operacion se realizo correctamente')
                return HttpResponseRedirect(reverse('usuarios:inicio'))

            else:
                messages.add_message(request, messages.ERROR, 'Su usuario no tiene los permisos de acceso a esta '
                                                              'seccion')
                return HttpResponseRedirect(reverse('usuarios:inicio'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

class GeneradorConceptos(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/consumosuscriptor.html'
    facturador = FacturadorConceptos

    def get(self, request):
        try:
            usuario = Usuario.objects.get(usuid=request.user.pk)
            idacueducto = usuario.IdAcueducto
            acueducto = Acueducto.objects.get(IdAcueducto=usuario.IdAcueducto)
            idtarifa = acueducto.IdTarifa
            tarifa = Tarifa.objects.get(IdTarifa=idtarifa.pk)
            aportefijo = tarifa.Valor
            recargo = tarifa.Recargo
            m3 = tarifa.m3
            valormetro = tarifa.valormetro
            fechaexp = (datetime.today())
            mes1 = fechaexp.month
            ciclos = Ciclo.objects.get(IdCiclo=mes1)
            mes = ciclos.Nombre
            ano = fechaexp.year
            tipousuario = True
            viviendas = Vivienda.objects.all()
            consumos = Consumos.objects.filter(mes=mes,ano=ano,IdAcueducto=idacueducto)
            conceptosfacturados = ConceptosFacturados.objects.filter(Estado='Pendiente',IdAcueducto=idacueducto)
            suspenciones = OrdenesTrabajo.objects.filter(Estado='Pendiente',IdAcueducto=idacueducto)
            facturas = Facturas.objects.filter(Estado='Emitida',IdAcueducto=idacueducto)
            matriculas = CobroMatricula.objects.filter(Estado='Pendiente',IdAcueducto=idacueducto)
            acuerdospago = AcuerdosPago.objects.filter(Estado='Pendiente',IdAcueducto=idacueducto)
            if tipousuario is True:
                for i in suspenciones:
                    cambio = OrdenesTrabajo.objects.get(IdOrden=i.IdOrden)
                    cambio.Estado = 'Anulada'
                    cambio.save()

                for i in facturas:
                    cambio = Facturas.objects.get(IdFactura=i.IdFactura)
                    cambio.Estado = 'Vencida'
                    cambio.save()

                for i in conceptosfacturados:
                    if i.Total >= TARIFASUSPENCION:
                        concepto = Conceptos(Tipo='Recargo', Observacion=str(i.Periodo) + ' - Valor vencido: - $' + str(i.Total), Estado='Sin facturar',
                                             Valor=recargo, IdVivienda=i.IdVivienda)
                        concepto.save()

                for i in conceptosfacturados:
                    if i.Total >= 1:
                        concepto1 = Conceptos(Tipo='Saldo P', Observacion=str(i.IdRegistro)+' - ' + str(i.Periodo) + ' - Saldo pendiente: - $' + str(i.Total),
                                             Estado='Sin facturar',Valor=i.Total, IdVivienda=i.IdVivienda)
                        concepto1.save()
                        conceptofac = ConceptosFacturados.objects.get(IdRegistro=i.IdRegistro)
                        conceptofac.Estado = 'Vencido'
                        conceptofac.save()

                for i in viviendas:
                    if i.EstadoServicio == 'Operativo' and i.TipoRecaudo =='Aporte fijo':
                        vivienda = Vivienda.objects.get(IdVivienda=i.IdVivienda)
                        concepto = Conceptos(Tipo='Aporte fijo', Observacion=mes, Estado='Sin facturar',
                                             Valor=aportefijo, IdVivienda=vivienda)
                        concepto.save()

                    if i.EstadoServicio == 'Operativo' and i.TipoRecaudo == 'Medicion':
                        vivienda = Vivienda.objects.get(IdVivienda=i.IdVivienda)
                        concepto = Conceptos(Tipo='Aporte fijo', Observacion=mes, Estado='Sin facturar',
                                             Valor=3000, IdVivienda=vivienda)
                        concepto.save()

                for i in consumos:
                    if int(i.Consumo) >= 1 and int(i.Consumo) <= 20:
                        resultado = int(valormetro) * int(i.Consumo)
                        concepto = Conceptos(Tipo='Basico',
                                             Observacion=mes + ' - Consumo m3: ' + str(i.Consumo),
                                             Estado='Sin facturar', Valor=resultado, IdVivienda=i.IdVivienda)
                        concepto.save()

                    elif int(i.Consumo) >= 21:
                        resultado = int(valormetro) * 20
                        consumo1 = 20
                        t2 = 350
                        concepto = Conceptos(Tipo='Basico',
                                             Observacion=mes + ' - Consumo m3: ' + str(consumo1),
                                             Estado='Sin facturar', Valor=resultado, IdVivienda=i.IdVivienda)
                        concepto.save()

                        sobrante = int(i.Consumo) - 20
                        resultado2 = t2 * sobrante
                        concepto2 = Conceptos(Tipo='Consumo complementario', Observacion=mes + ' - Consumo m3: ' + str(sobrante),
                                             Estado='Sin facturar', Valor=resultado2, IdVivienda=i.IdVivienda)
                        concepto2.save()

                for i in acuerdospago:
                    if int(i.CuotasPendientes) >=2:
                        editarcobro = AcuerdosPago.objects.get(IdVivienda=i.IdVivienda)
                        editarcobro.CuotasPendientes = int(i.CuotasPendientes) - 1
                        editarcobro.ValorPendiente = int(i.ValorPendiente) - int(i.Cuota)
                        editarcobro.save()
                        concepto = Conceptos(Tipo='Financiacion', Observacion=mes + ' - Cuota No: ' + i.CuotasPendientes,
                                             Estado='Sin facturar', Valor=i.Cuota, IdVivienda=i.IdVivienda)
                        concepto.save()

                    else:
                        editarcobro = AcuerdosPago.objects.get(IdVivienda=i.IdVivienda)
                        editarcobro.CuotasPendientes = int(i.CuotasPendientes) - 1
                        editarcobro.ValorPendiente = int(i.ValorPendiente) - int(i.Cuota)
                        editarcobro.Estado = 'Pago'
                        editarcobro.save()
                        concepto = Conceptos(Tipo='Financiacion', Observacion=mes + ' - Cuota No: ' + i.CuotasPendientes,
                                                 Estado='Sin facturar', Valor=i.Cuota, IdVivienda=i.IdVivienda)
                        concepto.save()

                ver = self.facturador()
                ejercutar = ver.get(request)
                return ejercutar

            else:
                messages.add_message(request, messages.ERROR, 'Su usuario no tiene los permisos de acceso a esta '
                                                              'seccion')
                return HttpResponseRedirect(reverse('usuarios:inicio'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

class GeneradorFacturas(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/generadorfacturas.html'

    def get(self, request):
        try:
            usuario = Usuario.objects.get(usuid=request.user.pk)
            idacueducto = usuario.IdAcueducto
            acueducto = Acueducto.objects.get(IdAcueducto=idacueducto)
            fechaexp = (datetime.today())
            mes1 = fechaexp.month
            ciclos = Ciclo.objects.get(IdCiclo=mes1)
            mes = ciclos.Nombre
            ano = fechaexp.year
            fechalimite = fechaexp + timedelta(days=8)
            facemi = Facturas.objects.filter(Estado=FE,IdAcueducto=idacueducto).count()

            viviendas = Vivienda.objects.filter(EstadoServicio='Operativo',IdAcueducto=idacueducto)|Vivienda.objects.filter(EstadoServicio='Suspendido',IdAcueducto=idacueducto)

            if facemi <=0:
                for i in viviendas:
                    conceptos = ConceptosFacturados.objects.filter(Estado='Pendiente', IdVivienda=i.IdVivienda).count()
                    if conceptos == 1:
                        conceptos2 = ConceptosFacturados.objects.get(Estado='Pendiente', IdVivienda=i.IdVivienda)
                        if conceptos2.SaldoAnterior >1:
                            valor = conceptos2.Total
                            vivienda = Vivienda.objects.get(IdVivienda=i.IdVivienda)
                            factura = Facturas(Estado='Emitida',IdVivienda=vivienda,periodofacturado=mes,FechaExpe=fechaexp,
                                               FechaLimite=fechalimite,facturasvencidas=1,IdCiclo=ciclos, Total=valor, IdAcueducto=acueducto)
                            factura.save()
                            relacion = FacturasConceptos(IdConcepto=conceptos2,IdFactura=factura)
                            relacion.save()

                        else:
                            valor = conceptos2.Total
                            vivienda = Vivienda.objects.get(IdVivienda=i.IdVivienda)
                            factura = Facturas(Estado='Emitida', IdVivienda=vivienda, periodofacturado=mes,
                                               FechaExpe=fechaexp, FechaLimite=fechalimite, facturasvencidas=0, IdCiclo=ciclos, Total=valor,
                                               IdAcueducto=acueducto)
                            factura.save()
                            relacion = FacturasConceptos(IdConcepto=conceptos2, IdFactura=factura)
                            relacion.save()

                messages.add_message(request, messages.ERROR, 'No se pudo generar la facturacion')
                return HttpResponseRedirect(reverse('usuarios:inicio'))

            else:
                messages.add_message(request, messages.ERROR,'No se pudo generar la facturacion')
                return HttpResponseRedirect(reverse('usuarios:inicio'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")


class ReporteConsumos(LoginRequiredMixin, View):
    login_url = '/'

    def get(self, request):
        usuario = Usuario.objects.get(usuid=request.user.pk)
        idacueducto = usuario.IdAcueducto
        consumos = Consumos.objects.filter(IdAcueducto=idacueducto)
        sfecha = (datetime.today())
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte - consumos"
        ws['A1'] = 'IdRegistro'
        ws['B1'] = 'IdMedidor'
        ws['C1'] = 'IdVivienda'
        ws['D1'] = 'Lecturaactual'
        ws['E1'] = 'Lecturaanterior'
        ws['F1'] = 'Consumo'
        ws['G1'] = 'promedio'
        ws['H1'] = 'observaciones'
        ws['I1'] = 'diasconsumo'
        ws['J1'] = 'ano'
        ws['K1'] = 'mes'
        ws['L1'] = 'Fecha'
        cont = 2
        for consu in consumos:
            ws.cell(row=cont, column=1).value = consu.IdRegistro
            ws.cell(row=cont, column=2).value = str(consu.IdMedidor)
            ws.cell(row=cont, column=3).value = str(consu.IdVivienda.pk)
            ws.cell(row=cont, column=4).value = consu.Lecturaactual
            ws.cell(row=cont, column=5).value = consu.Lecturaanterior
            ws.cell(row=cont, column=6).value = consu.Consumo
            ws.cell(row=cont, column=7).value = consu.promedio
            ws.cell(row=cont, column=8).value = consu.observaciones
            ws.cell(row=cont, column=9).value = consu.diasconsumo
            ws.cell(row=cont, column=10).value = consu.ano
            ws.cell(row=cont, column=11).value = consu.mes
            ws.cell(row=cont, column=12).value = consu.Fecha
            cont += 1

        archivo_predios = "Reporte - consumos " + str(sfecha) + ".xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(archivo_predios)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


class PlantillaMedicion(LoginRequiredMixin, View):
    login_url = '/'

    def get(self, request):
        usuario = Usuario.objects.get(usuid=request.user.pk)
        idacueducto = usuario.IdAcueducto
        consumos = Asignacion.objects.filter(IdAcueducto=idacueducto, Estado="Operativo")
        sfecha = (datetime.today())
        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla medicion"
        ws['A1'] = 'No'
        ws['B1'] = 'IdMedidor'
        ws['C1'] = 'Matricula'
        ws['D1'] = 'Direccion'
        ws['E1'] = 'Titular'
        ws['F1'] = 'LecturaActural'
        cont = 2
        suma = 1
        for consu in consumos:
            ws.cell(row=cont, column=1).value = suma
            ws.cell(row=cont, column=2).value = str(consu.IdMedidor)
            ws.cell(row=cont, column=3).value = str(consu.IdVivienda.pk)
            vivienda= Vivienda.objects.get(IdVivienda=consu.IdVivienda.pk)
            ws.cell(row=cont, column=4).value = str(vivienda.Direccion +' CS ' + vivienda.NumeroCasa)
            ws.cell(row=cont, column=5).value = str(vivienda.IdPropietario.Nombres + ' ' + vivienda.IdPropietario.Apellidos)
            cont += 1
            suma += 1

        archivo_predios = "Reporte - consumos .xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(archivo_predios)
        response['Content-Disposition'] = content
        wb.save(response)
        return response

class Varias(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/consumosuscriptor.html'

    def get(self, request):
        try:
            usuario = Usuario.objects.get(usuid=request.user.pk)
            acueducto = Acueducto.objects.get(IdAcueducto=usuario.IdAcueducto)
            idtarifa = acueducto.IdTarifa
            tarifa = Tarifa.objects.get(IdTarifa=idtarifa.pk)
            fechaexp = (datetime.today())
            mes1 = fechaexp.month
            ciclos = Ciclo.objects.get(IdCiclo=mes1)

            consumos = ConceptosFacturados.objects.filter(Estado='Pendiente')
            for i in consumos:
                if int(i.SaldoAnterior) >= 1:
                    vivienda = Vivienda.objects.get(IdVivienda=i.IdVivienda.pk)
                    orden = OrdenesTrabajo(Deuda=i.Total,Estado='Pendiente',TipoNovedad='Suspension',FechaEjecucion=fechaexp,
                                               usuario=usuario, IdVivienda=vivienda,IdAcueducto=acueducto)
                    orden.save()

                else:
                    pass

            messages.add_message(request, messages.INFO, 'La operacion se realizo correctamente')
            return HttpResponseRedirect(reverse('usuarios:inicio'))

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")

class MapaMedidores(LoginRequiredMixin, View):
    login_url = '/'
    template_name = 'usuarios/mapamedidores.html'

    def get(self, request):
        try:
            pasonivel = Vivienda.objects.filter(Direccion='Pasonivel Destapada')|Vivienda.objects.filter(Direccion='Pasonivel Viejo')

            return render(request, self.template_name,{'viviendas': pasonivel})

        except ObjectDoesNotExist:
            return render(request, "pages-404.html")